# -*- coding: utf-8 -*-
"""
商机拆表工具 - Flask 后端应用
功能：上传Excel、编辑分局人员映射、按分局拆分、下载结果
"""

import os
import json
import re
import shutil
import zipfile
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder='static', static_url_path='')
CORS(app)

# ============ 数据存储 ============
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
UPLOAD_DIR = os.path.join(DATA_DIR, 'uploads')
OUTPUT_DIR = os.path.join(DATA_DIR, 'output')
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

MAPPING_FILE = os.path.join(DATA_DIR, 'bureau_mapping.json')

# ============ 默认分局人员映射 ============
DEFAULT_MAPPING = {
    "工业能源政企分局": ["朱晨静", "肖智宇", "杜云帆", "屈容", "陈谦", "唐璐", "杜秋宇"],
    "国有平台政企分局": ["付登会", "蓝旭辉", "田小兰", "颜小琳", "易琴", "杨文", "王越"],
    "健康医疗政企分局": ["潘昱忻", "黎明", "先有为", "王虹丹"],
    "金融证券政企分局": ["贺贤珍", "曹岚", "梁曦宇", "杨彭萱", "张静", "郑文", "刘浩林", "杨真", "杨玮萍"],
    "软件科研政企分局": ["雷世豪", "陈燕", "杨璨宇", "李春江", "潘邓浩", "蒋尧莉", "李永婷", "蒲竑佚", "周建龙", "杨楚琪"],
    "新经济政企分局": ["曾理", "韩思萌", "罗维", "罗霞", "杨佩东", "刘文博", "戚新鹏"],
    "政法应急政企分局": ["王万辞", "徐杨", "袁进", "朱林", "袁新博", "张皓秋", "林佳俊", "黄天玉"],
    "政务政企分局": ["薛笑枫", "廖晓东", "钱宇煊", "吴倩", "杜成思", "黄振宇"],
    "高新孵化园智改数转服务局": ["董小凤", "邱国锋", "龙俊儒"],
    "高新天府生命科技园智改数转服务局": ["杨佳", "蒋建国", "王斯祺"],
    "高新天府软件园智改数转服务局": ["谢思宇", "黄微", "李选玉", "朱勇", "谢勇", "何琼", "高欢", "周莉", "李艺"],
    "金融城商客分局": ["姚尧", "王淑惠", "陈伟智", "曾宇嘉", "罗中伟", "张成铭", "刘荣"],
    "新川商客分局": ["靳扬", "姜春阳", "郑黎霞", "杨晋"],
    "天府新谷商客分局": ["李思锐", "彭倩", "黄燕", "刘鹏洋", "刘星月"],
    "新会展商客分局": ["何艳", "樊志林", "周滨", "蒋稚薇"],
    "天府国际商客分局": ["叶江", "蒋天佑", "杨茜", "李巧巧", "王辰雨", "廖华", "曾明全", "巫婷婷"],
    "环球商客分局": ["曾明", "许可", "钟小燕", "肖福洋", "冯麟霞", "王琴丽"],
    "大源商客分局": ["邱浩锋", "冯兰越", "冯特峰", "裴嘉轩", "何亚琪"],
    "肖芳商客分局": ["任登科", "贾小东", "梁润", "陈雪"],
    "府城商客分局": ["陈磊", "李若玉", "张小龙", "孙雯"],
    "连锁商客分局": ["杨凤翥", "肖帆", "赵娇", "高毛茅", "温有军"],
    "西信商客分局": ["杨力", "王宇", "任少杰", "周雨晴", "刘祖源", "胡文瀚", "赵川川"],
    "东苑商客分局": ["雷蕾", "吴文宪", "孙艺丹", "张宇魁", "聂海林", "陈健明"]
}


def load_mapping():
    """加载分局人员映射"""
    if os.path.exists(MAPPING_FILE):
        try:
            with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return DEFAULT_MAPPING.copy()


def save_mapping(mapping):
    """保存分局人员映射"""
    with open(MAPPING_FILE, 'w', encoding='utf-8') as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)


def clean_name(name):
    """清洗姓名，去除括号内的内容"""
    if not isinstance(name, str):
        return ""
    name = re.sub(r'\([^)]*\)', '', name)
    name = re.sub(r'\（[^）]*\）', '', name)
    return name.strip()


def copy_row_with_format(source_sheet, target_sheet, source_row, target_row, max_col):
    """复制行并保持格式"""
    for col in range(1, max_col + 1):
        source_cell = source_sheet.cell(row=source_row, column=col)
        target_cell = target_sheet.cell(row=target_row, column=col)
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell.font = source_cell.font.copy()
            target_cell.border = source_cell.border.copy()
            target_cell.fill = source_cell.fill.copy()
            target_cell.number_format = source_cell.number_format
            target_cell.protection = source_cell.protection.copy()
            target_cell.alignment = source_cell.alignment.copy()
        col_letter = get_column_letter(col)
        target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width


def copy_sheet_with_format(source_sheet, target_sheet, row_indices):
    """复制工作表并保持格式"""
    max_col = source_sheet.max_column
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width
    target_row = 1
    for source_row in row_indices:
        copy_row_with_format(source_sheet, target_sheet, source_row, target_row, max_col)
        if source_sheet.row_dimensions[source_row].height:
            target_sheet.row_dimensions[target_row].height = source_sheet.row_dimensions[source_row].height
        target_row += 1
    target_sheet.sheet_format = source_sheet.sheet_format
    target_sheet.page_setup = source_sheet.page_setup
    target_sheet.page_margins = source_sheet.page_margins


# ============ API 路由 ============

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')


@app.route('/api/mapping', methods=['GET'])
def get_mapping():
    """获取分局人员映射"""
    mapping = load_mapping()
    return jsonify(mapping)


@app.route('/api/mapping', methods=['POST'])
def save_mapping_api():
    """保存分局人员映射"""
    mapping = request.json
    if not mapping or not isinstance(mapping, dict):
        return jsonify({'error': '无效的映射数据'}), 400
    save_mapping(mapping)
    return jsonify({'message': '保存成功'})


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """上传Excel文件"""
    if 'file' not in request.files:
        return jsonify({'error': '没有找到上传文件'}), 400
    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '请上传Excel文件(.xlsx/.xls)'}), 400

    # 清理旧上传
    for f in os.listdir(UPLOAD_DIR):
        os.remove(os.path.join(UPLOAD_DIR, f))

    # 用安全文件名保存，避免中文路径问题
    import uuid
    ext = os.path.splitext(file.filename)[1]
    safe_filename = f"upload_{uuid.uuid4().hex[:8]}{ext}"
    original_filename = file.filename
    filepath = os.path.join(UPLOAD_DIR, safe_filename)
    file.save(filepath)

    # 读取表头信息
    try:
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value:
                headers.append(str(cell.value))
        total_rows = ws.max_row - 1 if ws.max_row else 0
        wb.close()

        # 同时用pandas读取客户经理列
        df = pd.read_excel(filepath, dtype=str)
        manager_col = None
        for col in df.columns:
            if '客户经理' in col:
                manager_col = col
                break
        managers = []
        if manager_col:
            managers = df[manager_col].dropna().apply(clean_name).unique().tolist()

        return jsonify({
            'filename': original_filename,
            'headers': headers,
            'total_rows': total_rows,
            'managers': managers
        })
    except Exception as e:
        return jsonify({'error': f'读取文件失败: {str(e)}'}), 500


@app.route('/api/split', methods=['POST'])
def split_file():
    """执行拆表操作"""
    data = request.json or {}
    mapping = data.get('mapping') or load_mapping()

    # 找上传文件
    uploaded = os.listdir(UPLOAD_DIR)
    if not uploaded:
        return jsonify({'error': '请先上传Excel文件'}), 400

    file_path = os.path.join(UPLOAD_DIR, uploaded[0])

    # 清理旧输出
    for f in os.listdir(OUTPUT_DIR):
        fp = os.path.join(OUTPUT_DIR, f)
        if os.path.isfile(fp):
            os.remove(fp)
        elif os.path.isdir(fp):
            shutil.rmtree(fp)

    current_date = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_folder = os.path.join(OUTPUT_DIR, f"分局拆分结果_{current_date}")
    os.makedirs(output_folder, exist_ok=True)

    try:
        source_wb = load_workbook(file_path)
        source_sheet = source_wb.active
    except Exception as e:
        return jsonify({'error': f'读取文件出错: {str(e)}'}), 500

    try:
        df = pd.read_excel(file_path, sheet_name=source_sheet.title, dtype=str)
    except Exception as e:
        return jsonify({'error': f'读取数据出错: {str(e)}'}), 500

    # 匹配
    bureau_rows = {bureau: [] for bureau in mapping.keys()}
    unmatched_rows = []
    header_row = 1
    matched_count = 0
    unmatched_count = 0
    unmatched_managers = set()

    for index in range(len(df)):
        excel_row = index + 2
        if '客户经理' in df.columns:
            manager_name_raw = df.iloc[index]['客户经理']
        elif 'AB1_客户经理' in df.columns:
            manager_name_raw = df.iloc[index]['AB1_客户经理']
        else:
            unmatched_rows.append(excel_row)
            unmatched_count += 1
            continue

        manager_name_clean = clean_name(manager_name_raw)
        matched = False
        for bureau, managers in mapping.items():
            if manager_name_clean in managers:
                bureau_rows[bureau].append(excel_row)
                matched_count += 1
                matched = True
                break

        if not matched:
            unmatched_rows.append(excel_row)
            unmatched_count += 1
            if manager_name_clean:
                unmatched_managers.add(manager_name_clean)

    # 为每个分局生成文件
    generated_files = []
    for bureau_name, row_indices in bureau_rows.items():
        if row_indices:
            target_wb = Workbook()
            target_sheet = target_wb.active
            all_rows = [header_row] + row_indices
            copy_sheet_with_format(source_sheet, target_sheet, all_rows)
            safe_name = re.sub(r'[\/\\:*?"<>|]', '_', bureau_name)
            output_file = os.path.join(output_folder, f"{safe_name}_{current_date}.xlsx")
            target_wb.save(output_file)
            generated_files.append({
                'bureau': bureau_name,
                'rows': len(row_indices),
                'filename': f"{safe_name}_{current_date}.xlsx"
            })

    # 汇总文件
    summary_file = os.path.join(output_folder, f"行业商机数据汇总_{current_date}.xlsx")
    summary_wb = Workbook()
    if 'Sheet' in summary_wb.sheetnames:
        del summary_wb['Sheet']
    for bureau_name, row_indices in bureau_rows.items():
        sheet_name = bureau_name[:31]
        target_sheet = summary_wb.create_sheet(title=sheet_name)
        if row_indices:
            all_rows = [header_row] + row_indices
            copy_sheet_with_format(source_sheet, target_sheet, all_rows)
    summary_wb.save(summary_file)
    generated_files.append({
        'bureau': '— 汇总文件 —',
        'rows': matched_count,
        'filename': f"行业商机数据汇总_{current_date}.xlsx"
    })

    # 未匹配名单
    if unmatched_rows:
        unmatched_wb = Workbook()
        unmatched_sheet = unmatched_wb.active
        all_rows = [header_row] + unmatched_rows
        copy_sheet_with_format(source_sheet, unmatched_sheet, all_rows)
        unmatched_file_path = os.path.join(output_folder, f"未匹配名单_{current_date}.xlsx")
        unmatched_wb.save(unmatched_file_path)
        generated_files.append({
            'bureau': '— 未匹配名单 —',
            'rows': unmatched_count,
            'filename': f"未匹配名单_{current_date}.xlsx"
        })

    # 打包为zip
    zip_name = f"分局拆分结果_{current_date}.zip"
    zip_path = os.path.join(OUTPUT_DIR, zip_name)
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for f in os.listdir(output_folder):
            zf.write(os.path.join(output_folder, f), f)

    source_wb.close()

    return jsonify({
        'message': '拆分完成',
        'matched': matched_count,
        'unmatched': unmatched_count,
        'files': generated_files,
        'unmatched_managers': sorted(unmatched_managers),
        'zip': zip_name,
        'output_folder': f"分局拆分结果_{current_date}"
    })


@app.route('/api/download/<path:filename>')
def download_file(filename):
    """下载文件"""
    return send_file(os.path.join(OUTPUT_DIR, filename), as_attachment=True)


@app.route('/api/download-folder/<path:folder>')
def download_folder(folder):
    """下载整个文件夹（zip）"""
    folder_path = os.path.join(OUTPUT_DIR, folder)
    if not os.path.isdir(folder_path):
        return jsonify({'error': '文件夹不存在'}), 404
    zip_name = f"{folder}.zip"
    zip_path = os.path.join(OUTPUT_DIR, zip_name)
    if not os.path.exists(zip_path):
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for f in os.listdir(folder_path):
                zf.write(os.path.join(folder_path, f), f)
    return send_file(zip_path, as_attachment=True)


@app.route('/api/reset-mapping', methods=['POST'])
def reset_mapping():
    """重置为默认映射"""
    save_mapping(DEFAULT_MAPPING.copy())
    return jsonify({'message': '已重置为默认映射'})


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5556, debug=True)
