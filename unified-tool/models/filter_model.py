# -*- coding: utf-8 -*-
import os
import re
import json
import shutil
import tempfile
import zipfile
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from config import (OUTPUT_DIR, DEFAULT_MAPPING, INDUSTRY_BUREAUS, COMMERCIAL_BUREAUS, DEFAULT_SPLIT_GROUPS)
from services.excel_service import clean_name, copy_sheet_with_format
from models.file_model import load_mapping


def split_filtered_data(file_bytes, filtered_indices, mapping, split_column, split_groups=None):
    tmp_in = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_in.write(file_bytes)
    tmp_in.close()

    for f in os.listdir(OUTPUT_DIR):
        fp = os.path.join(OUTPUT_DIR, f)
        if os.path.isfile(fp):
            os.remove(fp)
        elif os.path.isdir(fp):
            shutil.rmtree(fp)

    current_date = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_folder = os.path.join(OUTPUT_DIR, f"分局拆分结果_{current_date}")
    os.makedirs(output_folder, exist_ok=True)

    try:
        source_wb = load_workbook(tmp_in.name)
        source_sheet = source_wb.active
    except Exception as e:
        os.unlink(tmp_in.name)
        return {'ok': False, 'error': f'读取文件出错: {str(e)}'}

    try:
        df = pd.read_excel(tmp_in.name, sheet_name=source_sheet.title, dtype=str)
    except Exception as e:
        source_wb.close()
        os.unlink(tmp_in.name)
        return {'ok': False, 'error': f'读取数据出错: {str(e)}'}

    filtered_set = set(filtered_indices) if filtered_indices else set(range(len(df)))

    if not split_column:
        source_wb.close()
        os.unlink(tmp_in.name)
        return {'ok': False, 'error': '未指定拆分列'}
    if split_column not in df.columns:
        source_wb.close()
        os.unlink(tmp_in.name)
        return {'ok': False, 'error': f'拆分列 "{split_column}" 不存在于数据中'}

    bureau_rows = {bureau: [] for bureau in mapping.keys()}
    unmatched_rows = []
    header_row = 1
    matched_count = 0
    unmatched_count = 0
    unmatched_managers = set()

    for index in filtered_set:
        if index < 0 or index >= len(df):
            continue
        excel_row = index + 2

        manager_name_raw = df.iloc[index][split_column]
        manager_name_clean = clean_name(manager_name_raw)
        matched = False
        for bureau, managers in mapping.items():
            if manager_name_clean in managers:
                bureau_rows[bureau].append(excel_row)
                matched_count += 1
                matched = True
                break

        if not matched:
            unmatched_rows.append(excel_row)
            unmatched_count += 1
            if manager_name_clean:
                unmatched_managers.add(manager_name_clean)

    generated_files = []
    for bureau_name, row_indices in bureau_rows.items():
        if row_indices:
            target_wb = Workbook()
            target_sheet = target_wb.active
            all_rows = [header_row] + row_indices
            copy_sheet_with_format(source_sheet, target_sheet, all_rows)
            safe_name = re.sub(r'[\/\\:*?"<>|]', '_', bureau_name)
            output_file = os.path.join(output_folder, f"{safe_name}_{current_date}.xlsx")
            target_wb.save(output_file)
            generated_files.append({
                'bureau': bureau_name,
                'rows': len(row_indices),
                'filename': f"{safe_name}_{current_date}.xlsx"
            })

    def make_category_summary(category_name, category_bureaus):
        cat_wb = Workbook()
        if 'Sheet' in cat_wb.sheetnames:
            del cat_wb['Sheet']
        cat_rows = 0
        for bureau_name in category_bureaus:
            row_indices = bureau_rows.get(bureau_name, [])
            if row_indices:
                sheet_name = bureau_name[:31]
                target_sheet = cat_wb.create_sheet(title=sheet_name)
                all_rows = [header_row] + row_indices
                copy_sheet_with_format(source_sheet, target_sheet, all_rows)
                cat_rows += len(row_indices)
        if cat_rows:
            safe_cat = re.sub(r'[\/\\:*?"<>|]', '_', category_name)
            cat_file = os.path.join(output_folder, f"{safe_cat}_{current_date}.xlsx")
            cat_wb.save(cat_file)
            generated_files.append({
                'bureau': f'- {category_name} -',
                'rows': cat_rows,
                'filename': f"{safe_cat}_{current_date}.xlsx"
            })

    summary_file = os.path.join(output_folder, f"汇总数据_{current_date}.xlsx")
    summary_wb = Workbook()
    if 'Sheet' in summary_wb.sheetnames:
        del summary_wb['Sheet']
    for bureau_name, row_indices in bureau_rows.items():
        sheet_name = bureau_name[:31]
        target_sheet = summary_wb.create_sheet(title=sheet_name)
        if row_indices:
            all_rows = [header_row] + row_indices
            copy_sheet_with_format(source_sheet, target_sheet, all_rows)
    summary_wb.save(summary_file)
    generated_files.append({
        'bureau': '- 汇总文件 -',
        'rows': matched_count,
        'filename': f"汇总数据_{current_date}.xlsx"
    })

    # 按拆分组生成汇总文件（而非固定的行业/商业）
    if split_groups is None:
        split_groups = DEFAULT_SPLIT_GROUPS
    for group_name, group_bureaus in split_groups.items():
        make_category_summary(f"{group_name}数据汇总", group_bureaus)

    if unmatched_rows:
        unmatched_wb = Workbook()
        unmatched_sheet = unmatched_wb.active
        all_rows = [header_row] + unmatched_rows
        copy_sheet_with_format(source_sheet, unmatched_sheet, all_rows)
        unmatched_file_path = os.path.join(output_folder, f"未匹配名单_{current_date}.xlsx")
        unmatched_wb.save(unmatched_file_path)
        generated_files.append({
            'bureau': '- 未匹配名单 -',
            'rows': unmatched_count,
            'filename': f"未匹配名单_{current_date}.xlsx"
        })

    zip_name = f"分局拆分结果_{current_date}.zip"
    zip_path = os.path.join(OUTPUT_DIR, zip_name)
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for f in os.listdir(output_folder):
            zf.write(os.path.join(output_folder, f), f)

    source_wb.close()
    os.unlink(tmp_in.name)

    return {
        'ok': True,
        'message': '拆分完成',
        'matched': matched_count,
        'unmatched': unmatched_count,
        'totalFiltered': len(filtered_set),
        'files': generated_files,
        'unmatched_managers': sorted(unmatched_managers),
        'zip': zip_name,
        'output_folder': f"分局拆分结果_{current_date}"
    }
