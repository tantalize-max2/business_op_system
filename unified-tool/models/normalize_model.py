# -*- coding: utf-8 -*-
"""标准化模板数据存取层（纯 JSON CRUD）

业务逻辑（公式引擎、模板填充）已迁移至 services/normalize_service.py。
本模块仅负责标准化模板的持久化：列表、保存、读取、删除。
"""
import os
from datetime import datetime
from openpyxl import load_workbook
from config import NZ_TEMPLATES_DIR
from utils.storage import load_json, save_json, safe_filename


def nz_template_path(name):
    return os.path.join(NZ_TEMPLATES_DIR, f"{safe_filename(name)}.json")


def list_nz_templates():
    """列出所有标准化模板的摘要信息（按保存时间倒序）。"""
    templates = []
    if not os.path.exists(NZ_TEMPLATES_DIR):
        return templates
    for fname in os.listdir(NZ_TEMPLATES_DIR):
        if not fname.endswith('.json'):
            continue
        data = load_json(os.path.join(NZ_TEMPLATES_DIR, fname), default={})
        if not data:
            continue
        templates.append({
            'name': data.get('name', fname[:-5]),
            'savedAt': data.get('savedAt', 0),
            'sheetCount': data.get('sheetCount', 0)
        })
    templates.sort(key=lambda t: t.get('savedAt', 0), reverse=True)
    return templates


def save_nz_template(name, file_data):
    """保存标准化模板（含 base64 文件数据）。

    会校验 Excel 有效性并提取 sheet 数量作为元数据。
    """
    import base64
    import tempfile
    try:
        raw = base64.b64decode(file_data)
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp.write(raw)
        tmp.close()
        wb = load_workbook(tmp.name)
        sheet_count = len(wb.sheetnames)
        wb.close()
        os.unlink(tmp.name)
    except Exception as e:
        return {'ok': False, 'error': f'模板文件无效: {str(e)}'}

    template_data = {
        'name': name,
        'fileData': file_data,
        'savedAt': datetime.now().timestamp() * 1000,
        'sheetCount': sheet_count
    }
    save_json(nz_template_path(name), template_data)
    return {'ok': True, 'name': name}


def get_nz_template(name):
    """读取单个标准化模板完整数据，不存在返回 None。"""
    fpath = nz_template_path(name)
    if not os.path.exists(fpath):
        return None
    return load_json(fpath, default=None)


def delete_nz_template(name):
    """删除标准化模板，不存在返回 False。"""
    fpath = nz_template_path(name)
    if not os.path.exists(fpath):
        return False
    os.remove(fpath)
    return True
