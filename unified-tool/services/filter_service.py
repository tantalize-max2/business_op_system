# -*- coding: utf-8 -*-
"""拆分服务层 - 商机数据按分局拆分的核心业务逻辑

从原 models/filter_model.py 迁移。本模块不含数据持久化（拆分结果直接写文件），
纯粹负责：Excel 解析 → 按分局匹配 → 生成多个分局文件 + 汇总文件 → 打包 ZIP。

格式保真策略：源工作簿只加载一次，所有输出文件通过在同一工作簿中创建新 sheet、
临时移除原 sheet、保存、恢复原 sheet 的方式生成，确保完整继承 styles.xml。
"""
import os
import re
import shutil
import tempfile
import zipfile
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from config import (OUTPUT_DIR, DEFAULT_MAPPING, INDUSTRY_BUREAUS,
                    COMMERCIAL_BUREAUS, DEFAULT_SPLIT_GROUPS)
from services.excel_service import clean_name, copy_sheet_with_format
from models.file_model import load_mapping


def split_filtered_data(file_bytes, filtered_indices, mapping, split_column, split_groups=None):
    """按分局拆分过滤后的数据。

    Args:
        file_bytes: 源 Excel 文件字节
        filtered_indices: 已过滤的行索引列表（None 表示全部）
        mapping: {分局名: [客户经理列表]}
        split_column: 拆分依据列名
        split_groups: {组名: [分局列表]}，None 时用默认拆分组

    Returns:
        dict: 拆分结果，含 matched/unmatched/files/zip 等
    """
    tmp_in = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_in.write(file_bytes)
    tmp_in.close()

    # 清空输出目录
    for f in os.listdir(OUTPUT_DIR):
        fp = os.path.join(OUTPUT_DIR, f)
        if os.path.isfile(fp):
            os.remove(fp)
        elif os.path.isdir(fp):
            shutil.rmtree(fp)

    current_date = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_folder = os.path.join(OUTPUT_DIR, f"分局拆分结果_{current_date}")
    os.makedirs(output_folder, exist_ok=True)

    try:
        source_wb = load_workbook(tmp_in.name)
        source_sheet = source_wb.active
    except Exception as e:
        os.unlink(tmp_in.name)
        return {'ok': False, 'error': f'读取文件出错: {str(e)}'}

    try:
        df = pd.read_excel(tmp_in.name, sheet_name=source_sheet.title, dtype=str)
    except Exception as e:
        source_wb.close()
        os.unlink(tmp_in.name)
        return {'ok': False, 'error': f'读取数据出错: {str(e)}'}

    filtered_set = set(filtered_indices) if filtered_indices else set(range(len(df)))

    if not split_column:
        source_wb.close()
        os.unlink(tmp_in.name)
        return {'ok': False, 'error': '未指定拆分列'}
    if split_column not in df.columns:
        source_wb.close()
        os.unlink(tmp_in.name)
        return {'ok': False, 'error': f'拆分列 "{split_column}" 不存在于数据中'}

    source_sheet_title = source_sheet.title

    bureau_rows = {bureau: [] for bureau in mapping.keys()}
    unmatched_rows = []
    header_row = 1
    matched_count = 0
    unmatched_count = 0
    unmatched_managers = set()
    total_filtered = len(filtered_set)

    for idx, index in enumerate(filtered_set):
        if index < 0 or index >= len(df):
            continue
        excel_row = index + 2

        manager_name_raw = df.iloc[index][split_column]
        manager_name_clean = clean_name(manager_name_raw)
        matched = False
        for bureau, managers in mapping.items():
            if manager_name_clean in managers:
                bureau_rows[bureau].append(excel_row)
                matched_count += 1
                matched = True
                break

        if not matched:
            unmatched_rows.append(excel_row)
            unmatched_count += 1
            if manager_name_clean:
                unmatched_managers.add(manager_name_clean)

    generated_files = []

    # ========== 生成各分局文件 ==========
    for bureau_name, row_indices in bureau_rows.items():
        if row_indices:
            _save_single_sheet(source_wb, source_sheet, source_sheet_title,
                               [header_row] + row_indices, output_folder,
                               bureau_name, current_date, generated_files,
                               bureau_label=bureau_name, row_count=len(row_indices))

    # ========== 生成汇总文件（多 sheet） ==========
    _save_multi_sheet(source_wb, source_sheet, source_sheet_title,
                      output_folder, '汇总数据', current_date,
                      bureau_rows, header_row, generated_files,
                      file_label='- 汇总文件 -', row_count=matched_count)

    # ========== 生成分类汇总 ==========
    if split_groups is None:
        split_groups = DEFAULT_SPLIT_GROUPS
    for group_name, group_bureaus in split_groups.items():
        cat_rows_map = {}
        cat_total = 0
        for bureau_name in group_bureaus:
            ri = bureau_rows.get(bureau_name, [])
            if ri:
                cat_rows_map[bureau_name] = ri
                cat_total += len(ri)
        if cat_rows_map:
            _save_multi_sheet(source_wb, source_sheet, source_sheet_title,
                              output_folder, group_name, current_date,
                              cat_rows_map, header_row, generated_files,
                              file_label=f'- {group_name} -', row_count=cat_total)

    # ========== 生成未匹配名单 ==========
    if unmatched_rows:
        _save_single_sheet(source_wb, source_sheet, source_sheet_title,
                           [header_row] + unmatched_rows, output_folder,
                           '未匹配名单', current_date, generated_files,
                           bureau_label='- 未匹配名单 -', row_count=unmatched_count)

    source_wb.close()

    # 打包 ZIP
    zip_name = f"分局拆分结果_{current_date}.zip"
    zip_path = os.path.join(OUTPUT_DIR, zip_name)
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for f in os.listdir(output_folder):
            zf.write(os.path.join(output_folder, f), f)

    os.unlink(tmp_in.name)

    return {
        'ok': True,
        'message': '拆分完成',
        'matched': matched_count,
        'unmatched': unmatched_count,
        'totalFiltered': len(filtered_set),
        'files': generated_files,
        'unmatched_managers': sorted(unmatched_managers),
        'zip': zip_name,
        'output_folder': f"分局拆分结果_{current_date}"
    }


def _save_single_sheet(source_wb, source_sheet, source_sheet_title, all_rows,
                       output_folder, name_prefix, current_date, generated_files,
                       bureau_label, row_count):
    """在源工作簿中创建新 sheet、复制行、临时移除原 sheet、保存、恢复。

    格式完全继承自源工作簿的 styles.xml / theme。
    """
    new_sheet = source_wb.create_sheet(title='Sheet')
    copy_sheet_with_format(source_sheet, new_sheet, all_rows)

    # 临时从工作簿移除源 sheet（保留对象引用以便恢复）
    source_wb._sheets.remove(source_sheet)
    try:
        safe_name = re.sub(r'[\/\\:*?"<>|]', '_', name_prefix)
        output_file = os.path.join(output_folder, f"{safe_name}_{current_date}.xlsx")
        source_wb.save(output_file)
    finally:
        # 恢复源 sheet 到首位，移除临时 sheet
        source_wb._sheets.remove(new_sheet)
        source_wb._sheets.insert(0, source_sheet)

    generated_files.append({
        'bureau': bureau_label,
        'rows': row_count,
        'filename': f"{safe_name}_{current_date}.xlsx"
    })


def _save_multi_sheet(source_wb, source_sheet, source_sheet_title,
                      output_folder, name_prefix, current_date,
                      rows_map, header_row, generated_files,
                      file_label, row_count):
    """生成多 sheet 汇总文件。

    在源工作簿中创建多个新 sheet，临时移除原 sheet，保存，恢复。
    """
    temp_sheets = []
    for sheet_name, row_indices in rows_map.items():
        if row_indices:
            new_sheet = source_wb.create_sheet(title=sheet_name[:31])
            copy_sheet_with_format(source_sheet, new_sheet, [header_row] + row_indices)
            temp_sheets.append(new_sheet)

    if not temp_sheets:
        return

    # 临时移除源 sheet
    source_wb._sheets.remove(source_sheet)
    try:
        safe_name = re.sub(r'[\/\\:*?"<>|]', '_', name_prefix)
        output_file = os.path.join(output_folder, f"{safe_name}_{current_date}.xlsx")
        source_wb.save(output_file)
    finally:
        # 移除所有临时 sheet，恢复源 sheet
        for ts in temp_sheets:
            source_wb._sheets.remove(ts)
        source_wb._sheets.insert(0, source_sheet)

    generated_files.append({
        'bureau': file_label,
        'rows': row_count,
        'filename': f"{safe_name}_{current_date}.xlsx"
    })
