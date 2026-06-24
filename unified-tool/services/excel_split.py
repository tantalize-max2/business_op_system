# -*- coding: utf-8 -*-
"""Excel 格式保真拆分工具。

核心策略：绕过 openpyxl 的样式重写，直接操作 xlsx 内部的 XML：
- 读取源 xlsx 的 sheet XML，按行过滤
- 重新编号行号，生成新的 sheet XML
- 保留原始 styles.xml / theme / sharedStrings.xml 等全部不动
- 输出文件与源文件格式 100% 一致

样式索引（s 属性）和共享字符串索引（t="s" + v 值）直接引用原始文件，
无需任何转换。
"""
import os
import re
import io
import zipfile
import xml.etree.ElementTree as ET
from openpyxl.utils import get_column_letter


def clean_name(name):
    if not isinstance(name, str):
        return ""
    name = re.sub(r'\([^)]*\)', '', name)
    name = re.sub(r'\（[^）]*\）', '', name)
    parts = re.split(r'[,，、;；\s]+', name)
    parts = [p.strip() for p in parts if p.strip()]
    return parts[0] if parts else name.strip()


def split_xlsx_fidelity(source_bytes, bureau_row_map, header_xml, max_col,
                        output_folder, current_date, split_groups,
                        unmatched_rows, matched_count, unmatched_count):
    """直接操作 XML 实现格式 100% 保真的拆分。

    Args:
        source_bytes: 源 xlsx 文件字节
        bureau_row_map: {分局名: [源Excel行号列表]}
        header_xml: 表头行的原始 XML 字符串（<row r="1">...</row>）
        max_col: 最大列数
        output_folder: 输出目录
        current_date: 日期时间字符串
        split_groups: 拆分组配置
        unmatched_rows: 未匹配行号列表
        matched_count: 匹配总数
        unmatched_count: 未匹配总数

    Returns:
        list: 生成的文件信息
    """
    # 将源文件读入内存 zip
    source_zip = zipfile.ZipFile(io.BytesIO(source_bytes), 'r')
    source_files = {name: source_zip.read(name) for name in source_zip.namelist()}
    source_zip.close()

    # 提取 sheet XML 的非数据部分（前缀和后缀）
    sheet_xml_name = _find_sheet_xml_name(source_files)
    sheet_content = source_files[sheet_xml_name].decode('utf-8')

    # 分离 sheet XML: 前缀 + sheetData + 后缀
    sd_start = sheet_content.find('<sheetData>')
    sd_close = sheet_content.find('</sheetData>')
    xml_prefix = sheet_content[:sd_start + len('<sheetData>')]
    xml_suffix = '</sheetData>' + sheet_content[sd_close + len('</sheetData>'):]
    all_rows_xml = sheet_content[sd_start + len('<sheetData>'):sd_close]

    # 解析所有行，建立行号 -> XML 的映射
    row_map = _parse_rows(all_rows_xml)

    generated_files = []

    # ========== 生成各分局文件 ==========
    for bureau_name, row_numbers in bureau_row_map.items():
        if not row_numbers:
            continue
        # 构建新行列表：表头(行1) + 数据行
        new_rows = [1] + row_numbers
        new_sheet_xml = _build_sheet_xml(xml_prefix, xml_suffix, row_map, new_rows)

        safe_name = re.sub(r'[\/\\:*?"<>|]', '_', bureau_name)
        output_file = os.path.join(output_folder, f"{safe_name}_{current_date}.xlsx")
        _write_xlsx(source_files, sheet_xml_name, new_sheet_xml, output_file)
        generated_files.append({
            'bureau': bureau_name,
            'rows': len(row_numbers),
            'filename': f"{safe_name}_{current_date}.xlsx"
        })

    # ========== 生成汇总文件 ==========
    summary_rows_map = {}
    total_summary_rows = 0
    for bureau_name, row_numbers in bureau_row_map.items():
        if row_numbers:
            summary_rows_map[bureau_name] = row_numbers
            total_summary_rows += len(row_numbers)

    summary_file = os.path.join(output_folder, f"汇总数据_{current_date}.xlsx")
    _write_multi_sheet_xlsx(source_files, sheet_xml_name, row_map,
                            summary_rows_map, summary_file, xml_prefix, xml_suffix)
    generated_files.append({
        'bureau': '- 汇总文件 -',
        'rows': matched_count,
        'filename': f"汇总数据_{current_date}.xlsx"
    })

    # ========== 生成分类汇总 ==========
    for group_name, group_bureaus in split_groups.items():
        cat_rows_map = {}
        cat_total = 0
        for bn in group_bureaus:
            rn = bureau_row_map.get(bn, [])
            if rn:
                cat_rows_map[bn] = rn
                cat_total += len(rn)
        if cat_rows_map:
            safe_cat = re.sub(r'[\/\\:*?"<>|]', '_', group_name)
            cat_file = os.path.join(output_folder, f"{safe_cat}_{current_date}.xlsx")
            _write_multi_sheet_xlsx(source_files, sheet_xml_name, row_map,
                                    cat_rows_map, cat_file, xml_prefix, xml_suffix)
            generated_files.append({
                'bureau': f'- {group_name} -',
                'rows': cat_total,
                'filename': f"{safe_cat}_{current_date}.xlsx"
            })

    # ========== 生成未匹配名单 ==========
    if unmatched_rows:
        new_rows = [1] + unmatched_rows
        new_sheet_xml = _build_sheet_xml(xml_prefix, xml_suffix, row_map, new_rows)
        unmatched_file = os.path.join(output_folder, f"未匹配名单_{current_date}.xlsx")
        _write_xlsx(source_files, sheet_xml_name, new_sheet_xml, unmatched_file)
        generated_files.append({
            'bureau': '- 未匹配名单 -',
            'rows': unmatched_count,
            'filename': f"未匹配名单_{current_date}.xlsx"
        })

    return generated_files


def _find_sheet_xml_name(source_files):
    """找到第一个 worksheet 的 XML 文件名。"""
    for name in source_files:
        if name.startswith('xl/worksheets/sheet') and name.endswith('.xml'):
            return name
    return 'xl/worksheets/sheet1.xml'


def _parse_rows(all_rows_xml):
    """解析 sheetData XML，建立 {行号: 行XML字符串} 映射。

    使用正则提取每个 <row>...</row> 块。
    """
    row_map = {}
    # 匹配 <row r="数字">...</row>
    pattern = re.compile(r'<row r="(\d+)"[^>]*>.*?</row>', re.DOTALL)
    for match in pattern.finditer(all_rows_xml):
        row_num = int(match.group(1))
        row_map[row_num] = match.group(0)
    return row_map


def _renumber_row_xml(row_xml, old_num, new_num, max_col):
    """重新编号行的 r 属性和其中所有 cell 的 r 属性。"""
    # 替换行号: <row r="old" -> <row r="new"
    result = row_xml.replace(f'<row r="{old_num}"', f'<row r="{new_num}"', 1)

    # 替换每个 cell 的行号: r="Aold" -> r="Anew", r="Bold" -> r="Bnew"
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        result = result.replace(f'r="{col_letter}{old_num}"', f'r="{col_letter}{new_num}"')

    return result


def _build_sheet_xml(xml_prefix, xml_suffix, row_map, row_numbers):
    """构建单 sheet 的 XML。

    Args:
        xml_prefix: sheetData 开始标签之前的内容（含 <sheetData>）
        xml_suffix: sheetData 结束标签之后的内容
        row_map: {行号: 行XML}
        row_numbers: 要包含的源行号列表（已含表头行1）

    Returns:
        完整的 sheet XML 字符串
    """
    # 找最大列数
    max_col = 32  # 默认值，后续从行数据中推断

    parts = []
    for new_idx, src_row_num in enumerate(row_numbers, start=1):
        if src_row_num in row_map:
            # 推断最大列数
            row_xml = row_map[src_row_num]
            cols_in_row = re.findall(r'r="([A-Z]+)\d+"', row_xml)
            for c in cols_in_row:
                col_num = 0
                for ch in c:
                    col_num = col_num * 26 + (ord(ch) - ord('A') + 1)
                if col_num > max_col:
                    max_col = col_num
            parts.append(_renumber_row_xml(row_xml, src_row_num, new_idx, max_col))

    rows_xml = ''.join(parts)

    # 更新 <dimension> 为实际数据范围。原始前缀里的 dimension 仍是源文件的完整范围
    # （如 A1:AF1457），若不更新，拆分后的小文件在 WPS/Excel 中“已用区域”、
    # Ctrl+End 跳转、滚动条与打印范围都会沿用原始大范围，表现为“格式/范围与原文件不同”。
    written_rows = len(parts)
    if written_rows > 0:
        last_col_letter = get_column_letter(max_col)
        new_dim = '<dimension ref="A1:%s%d"/>' % (last_col_letter, written_rows)
        xml_prefix = re.sub(r'<dimension\s+ref="[^"]*"\s*/>',
                            new_dim, xml_prefix, count=1)

    return xml_prefix + rows_xml + xml_suffix


def _write_xlsx(source_files, sheet_xml_name, new_sheet_xml, output_path):
    """写入单 sheet xlsx：复制源文件所有内容，只替换 sheet XML。"""
    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for name, data in source_files.items():
            if name == sheet_xml_name:
                zf.writestr(name, new_sheet_xml.encode('utf-8'))
            else:
                zf.writestr(name, data)


def _write_multi_sheet_xlsx(source_files, sheet_xml_name, row_map,
                            rows_map, output_path, xml_prefix, xml_suffix):
    """写入多 sheet xlsx：每个分局一个 sheet。

    需要重写 workbook.xml（<sheets>）、workbook.xml.rels、[Content_Types].xml 来注册多个 sheet。
    关键：Relationship / Override 的属性值（Type / ContentType 是带 / 的 URL）里含 '/'，
    不能用 [^/] 截断，否则会匹配不到、漏掉 styles/sharedStrings 等关系，或留下重复 Override，
    导致 WPS/Excel 打开时报“内容有问题、需要修复”。
    """
    bureau_names = list(rows_map.keys())
    n = len(bureau_names)

    sheet_content = source_files[sheet_xml_name].decode('utf-8')
    sd_start = sheet_content.find('<sheetData>')
    common_prefix = sheet_content[:sd_start + len('<sheetData>')]

    WS_TYPE = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet'
    WS_CT = 'application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml'

    # ---- 1. 重写 workbook.xml.rels ----
    # 保留所有非 worksheet 关系（sharedStrings/styles/theme 等，维持其原始 rId），
    # 工作表关系重新分配 rId（接在已用 rId 之后），避免 rId 冲突。
    orig_rels = source_files.get('xl/_rels/workbook.xml.rels', b'').decode('utf-8')
    # [^>]*? 可正确跨过含 '/' 的 URL，匹配到自闭合的 />
    all_rel_xmls = re.findall(r'<Relationship\b[^>]*?/>', orig_rels)
    non_ws_rels = [r for r in all_rel_xmls if '/worksheet' not in r]
    used_nums = [int(m) for m in re.findall(r'Id="rId(\d+)"', ''.join(non_ws_rels))]
    start = max(used_nums) + 1 if used_nums else 1
    ws_rids = [f'rId{start + i}' for i in range(n)]
    new_ws_rels = ''.join(
        f'<Relationship Id="{ws_rids[i]}" Type="{WS_TYPE}" Target="worksheets/sheet{i + 1}.xml"/>'
        for i in range(n))
    ns_match = re.search(r'<Relationships\b[^>]*>', orig_rels)
    rels_root = ns_match.group(0) if ns_match else '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
    full_rels_xml = rels_root + ''.join(non_ws_rels) + new_ws_rels + '</Relationships>'

    # ---- 2. 重写 workbook.xml 的 <sheets>（引用新的工作表 rId）----
    new_sheet_tags = []
    for i, bn in enumerate(bureau_names):
        safe = re.sub(r'[\/\\:*?"<>|]', '_', bn)[:28]
        new_sheet_tags.append(f'<sheet name="{safe}" sheetId="{i + 1}" state="visible" r:id="{ws_rids[i]}"/>')
    new_sheets_block = '<sheets>' + ''.join(new_sheet_tags) + '</sheets>'

    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for name, data in source_files.items():
            if name == sheet_xml_name:
                # 第一个分局写入 sheet1.xml
                first_rows = [1] + rows_map[bureau_names[0]]
                new_xml = _build_sheet_xml(common_prefix, xml_suffix, row_map, first_rows)
                zf.writestr(name, new_xml.encode('utf-8'))
            elif name == 'xl/workbook.xml':
                wb_xml = re.sub(r'<sheets>.*?</sheets>', new_sheets_block,
                                data.decode('utf-8'), count=1, flags=re.DOTALL)
                zf.writestr(name, wb_xml.encode('utf-8'))
            elif name == 'xl/_rels/workbook.xml.rels':
                zf.writestr(name, full_rels_xml.encode('utf-8'))
            elif name.startswith('xl/worksheets/') and name.endswith('.xml') and name != sheet_xml_name:
                pass  # 源里其他工作表（本例无），跳过
            elif name.startswith('xl/worksheets/_rels/') and name != 'xl/worksheets/_rels/sheet1.xml.rels':
                pass  # 跳过，下方统一复制 sheet1 的 rels
            elif name == '[Content_Types].xml':
                ct_xml = data.decode('utf-8')
                # 稳健移除原有 worksheet Override（属性顺序不限、ContentType 含 '/'）
                ct_xml = re.sub(r'<Override\b[^>]*?/>',
                                lambda m: '' if re.search(r'PartName="/xl/worksheets/sheet\d+\.xml"', m.group(0)) else m.group(0),
                                ct_xml)
                overrides = ''.join(f'<Override PartName="/xl/worksheets/sheet{i + 1}.xml" ContentType="{WS_CT}"/>'
                                    for i in range(n))
                ct_xml = ct_xml.replace('</Types>', overrides + '</Types>')
                zf.writestr(name, ct_xml.encode('utf-8'))
            else:
                zf.writestr(name, data)

        # 写第 2..N 个工作表文件
        for i in range(1, n):
            new_rows = [1] + rows_map[bureau_names[i]]
            new_xml = _build_sheet_xml(common_prefix, xml_suffix, row_map, new_rows)
            zf.writestr(f'xl/worksheets/sheet{i + 1}.xml', new_xml.encode('utf-8'))
            # 复制 sheet1 的 rels（图片/绘图等引用）
            base_rels_name = 'xl/worksheets/_rels/sheet1.xml.rels'
            if base_rels_name in source_files:
                zf.writestr(f'xl/worksheets/_rels/sheet{i + 1}.xml.rels',
                            source_files[base_rels_name])
