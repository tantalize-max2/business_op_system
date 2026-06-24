# -*- coding: utf-8 -*-
import re
from copy import copy
from openpyxl.utils import get_column_letter, range_boundaries


def clean_name(name):
    if not isinstance(name, str):
        return ""
    name = re.sub(r'\([^)]*\)', '', name)
    name = re.sub(r'\（[^）]*\）', '', name)
    parts = re.split(r'[,，、;；\s]+', name)
    parts = [p.strip() for p in parts if p.strip()]
    return parts[0] if parts else name.strip()


def copy_row_with_format(source_sheet, target_sheet, source_row, target_row, max_col):
    for col in range(1, max_col + 1):
        source_cell = source_sheet.cell(row=source_row, column=col)
        target_cell = target_sheet.cell(row=target_row, column=col)
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
        col_letter = get_column_letter(col)
        src_dim = source_sheet.column_dimensions.get(col_letter)
        if src_dim and src_dim.width:
            target_sheet.column_dimensions[col_letter].width = src_dim.width


def copy_sheet_with_format(source_sheet, target_sheet, row_indices):
    """将源表中指定行复制到目标表，保留完整格式。"""
    max_col = source_sheet.max_column

    # 1. 列宽（含自定义样式）
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        src_dim = source_sheet.column_dimensions.get(col_letter)
        if src_dim:
            tgt_dim = target_sheet.column_dimensions[col_letter]
            if src_dim.width:
                tgt_dim.width = src_dim.width
            if src_dim.hidden:
                tgt_dim.hidden = True

    # 2. 复制行数据 + 行级格式
    target_row = 1
    for source_row in row_indices:
        copy_row_with_format(source_sheet, target_sheet, source_row, target_row, max_col)
        src_rd = source_sheet.row_dimensions.get(source_row)
        if src_rd:
            tgt_rd = target_sheet.row_dimensions[target_row]
            if src_rd.height:
                tgt_rd.height = src_rd.height
            if src_rd.hidden:
                tgt_rd.hidden = True
        target_row += 1

    # 3. 合并单元格 - 按行号映射到新位置
    _copy_merged_cells(source_sheet, target_sheet, row_indices)

    # 4. 冻结窗格（仅复制表头行相关的冻结）
    if source_sheet.freeze_panes:
        fp = source_sheet.freeze_panes
        # 如果冻结位置在复制范围之内，则映射到新位置
        if isinstance(fp, str):
            boundaries_obj = range_boundaries(fp)
            freeze_col = boundaries_obj[0]
            freeze_row = boundaries_obj[1]
        else:
            freeze_col = fp.col_idx
            freeze_row = fp.row

        # 查找冻结行在新表中的位置
        new_freeze_row = None
        for new_idx, src_row in enumerate(row_indices, start=1):
            if src_row >= freeze_row:
                new_freeze_row = new_idx
                break
        if new_freeze_row is not None and new_freeze_row > 1:
            target_sheet.freeze_panes = target_sheet.cell(row=new_freeze_row, column=freeze_col)

    # 5. 自动筛选 - 映射行号
    if source_sheet.auto_filter and source_sheet.auto_filter.ref:
        auto_ref = source_sheet.auto_filter.ref
        # 仅在数据至少1行时设置筛选
        if target_row > 1:
            target_sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{target_row - 1}"

    # 6. 页面设置
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.page_setup = copy(source_sheet.page_setup)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)

    # 7. 条件格式 - 映射行号范围
    _copy_conditional_formatting(source_sheet, target_sheet, row_indices)

    # 8. 数据验证 - 映射行号范围
    _copy_data_validations(source_sheet, target_sheet, row_indices)


def _copy_merged_cells(source_sheet, target_sheet, row_indices):
    """复制合并单元格，按行号映射到新位置。"""
    if not source_sheet.merged_cells:
        return

    # 建立源行号 -> 目标行号映射
    src_to_tgt = {}
    for tgt_idx, src_row in enumerate(row_indices, start=1):
        src_to_tgt[src_row] = tgt_idx

    for merged_range in source_sheet.merged_cells.ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col

        # 合并区域的所有行都必须在复制范围内
        new_min_row = src_to_tgt.get(min_row)
        new_max_row = src_to_tgt.get(max_row)

        if new_min_row is not None and new_max_row is not None:
            # 检查中间行也都存在
            all_present = all(r in src_to_tgt for r in range(min_row, max_row + 1))
            if all_present:
                new_range = f"{get_column_letter(min_col)}{new_min_row}:{get_column_letter(max_col)}{new_max_row}"
                target_sheet.merge_cells(new_range)

    # 复制合并单元格的样式（值已在行复制时写入）
    for merged_range in source_sheet.merged_cells.ranges:
        min_row = merged_range.min_row
        min_col = merged_range.min_col
        new_min_row = src_to_tgt.get(min_row)
        if new_min_row is not None:
            src_cell = source_sheet.cell(row=min_row, column=min_col)
            tgt_cell = target_sheet.cell(row=new_min_row, column=min_col)
            if src_cell.has_style:
                tgt_cell.font = copy(src_cell.font)
                tgt_cell.border = copy(src_cell.border)
                tgt_cell.fill = copy(src_cell.fill)
                tgt_cell.alignment = copy(src_cell.alignment)
                tgt_cell.number_format = src_cell.number_format


def _copy_conditional_formatting(source_sheet, target_sheet, row_indices):
    """复制条件格式，映射行号范围。"""
    if not hasattr(source_sheet, 'conditional_formatting') or not source_sheet.conditional_formatting:
        return

    src_to_tgt = {}
    for tgt_idx, src_row in enumerate(row_indices, start=1):
        src_to_tgt[src_row] = tgt_idx

    for cf in source_sheet.conditional_formatting:
        for rule_range in cf.cells.ranges:
            min_row = rule_range.min_row
            max_row = rule_range.max_row
            min_col = rule_range.min_col
            max_col = rule_range.max_col

            new_min_row = src_to_tgt.get(min_row)
            new_max_row = src_to_tgt.get(max_row)
            if new_min_row is None or new_max_row is None:
                continue

            # 检查范围行是否都在映射中
            all_present = all(r in src_to_tgt for r in range(min_row, max_row + 1))
            if not all_present:
                continue

            for rule in cf.rules:
                target_sheet.conditional_formatting.add(
                    f"{get_column_letter(min_col)}{new_min_row}:{get_column_letter(max_col)}{new_max_row}",
                    copy(rule)
                )


def _copy_data_validations(source_sheet, target_sheet, row_indices):
    """复制数据验证，映射行号范围。"""
    if not source_sheet.data_validations or not source_sheet.data_validations.dataValidation:
        return

    src_to_tgt = {}
    for tgt_idx, src_row in enumerate(row_indices, start=1):
        src_to_tgt[src_row] = tgt_idx

    for dv in source_sheet.data_validations.dataValidation:
        try:
            sqref = dv.sqref
            if isinstance(sqref, str):
                # 简单处理：尝试映射每个单元格范围
                parts = sqref.split()
                new_parts = []
                for part in parts:
                    if ':' in part:
                        start_ref, end_ref = part.split(':')
                        # 解析范围
                        boundaries_obj = range_boundaries(part)
                        min_col_b, min_row_b, max_col_b, max_row_b = boundaries_obj
                        new_min_row = src_to_tgt.get(min_row_b)
                        new_max_row = src_to_tgt.get(max_row_b)
                        if new_min_row and new_max_row:
                            all_present = all(r in src_to_tgt for r in range(min_row_b, max_row_b + 1))
                            if all_present:
                                new_parts.append(
                                    f"{get_column_letter(min_col_b)}{new_min_row}:"
                                    f"{get_column_letter(max_col_b)}{new_max_row}"
                                )
                    else:
                        # 单个单元格引用
                        boundaries_obj = range_boundaries(part)
                        col_b, row_b = boundaries_obj[0], boundaries_obj[1]
                        new_row = src_to_tgt.get(row_b)
                        if new_row:
                            new_parts.append(f"{get_column_letter(col_b)}{new_row}")

                if new_parts:
                    new_dv = copy(dv)
                    new_dv.sqref = ' '.join(new_parts)
                    target_sheet.data_validations.add(new_dv)
        except Exception:
            pass


def create_workbook_from_source(source_wb):
    """基于源工作簿创建新工作簿，继承全局样式（主题、字体等）。

    返回 (new_wb, default_sheet) — default_sheet 需要由调用方决定是否保留。
    """
    from openpyxl import Workbook
    wb = Workbook()
    # 复制主题（确保颜色/字体一致）
    if hasattr(source_wb, 'theme') and source_wb.theme:
        wb.theme = source_wb.theme
    # 复制属性
    if hasattr(source_wb, 'properties') and source_wb.properties:
        wb.properties = copy(source_wb.properties)
    # 复制样式名映射
    if hasattr(source_wb, '_named_styles'):
        for style_name in source_wb.named_styles:
            if style_name not in wb.named_styles:
                try:
                    wb._named_styles[style_name] = copy(source_wb._named_styles[style_name])
                except Exception:
                    pass
    return wb, wb.active
