# -*- coding: utf-8 -*-
"""标准化服务层 - 公式引擎与模板填充的业务逻辑

从原 models/normalize_model.py 迁移。本模块负责：
- {{N:L1:Entry/Metric}} 公式解析
- Excel 单元格引用解析（含循环引用检测）
- SUM/AVG 函数求值
- = 开头表达式的安全求值（使用 utils.safe_eval）
- 模板填充（cell_edits / cell_formats 应用 + 公式批量计算）
"""
import os
import re
import tempfile
import math
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from utils.safe_eval import safe_eval_arithmetic


_formula_pattern = re.compile(r'\{\{(.+?)\}\}')
_cell_ref_pattern = re.compile(r'\b([A-Z]{1,3})(\d{1,5})\b')
_func_pattern = re.compile(r'\b(SUM|AVG)\s*\(([^)]+)\)', re.IGNORECASE)


# ========== 公式解析（{{...}}） ==========

def nz_resolve_formula_str(formula_str, stats_data):
    m = re.match(r'^\{\{(.+?)\}\}$', formula_str.strip())
    if not m:
        m = re.search(r'\{\{(.+?)\}\}', formula_str.strip())
        if not m:
            return {'ok': False, 'value': '格式错误'}
    inner = m.group(1).strip()
    slash_idx = inner.rfind('/')
    if slash_idx < 0:
        return {'ok': False, 'value': '格式错误(无/指标)'}
    metric = inner[slash_idx + 1:].strip()
    path = inner[:slash_idx].strip()
    parts = path.split(':')
    parts = [p.strip() for p in parts]

    file_idx = None
    try:
        file_idx = int(parts[0])
    except (ValueError, IndexError):
        return {'ok': False, 'value': '文件序号无效'}

    if file_idx < 1:
        return {'ok': False, 'value': '文件序号无效'}

    fi_str = str(file_idx)
    if fi_str not in stats_data:
        return {'ok': False, 'value': '未匹配(文件不存在)'}

    fd = stats_data[fi_str]
    entries = fd.get('entries', [])
    total = fd.get('total', {})
    sum_col = fd.get('sumCol', '')

    l1, col = '', ''
    entry_name = ''

    if len(parts) == 2 and parts[1] == '总合计':
        return _nz_resolve_metric(total, '', metric, sum_col, None, None, fd)

    if len(parts) == 2:
        entry_name = parts[1]
    elif len(parts) == 3:
        l1 = parts[1]
        entry_name = parts[2]
    elif len(parts) == 4:
        l1 = parts[1]
        entry_name = parts[2]
        col = parts[3]
    elif len(parts) >= 5:
        l1 = parts[1]
        remaining = parts[2:]
        if remaining and '.' in remaining[-1]:
            col = remaining[-1]
            entry_name = ' · '.join(remaining[:-1])
        else:
            entry_name = ' · '.join(remaining)

    ac_col, ac_val = None, None
    if col and '.' in col:
        dp = col.split('.', 1)
        ac_col, ac_val = dp[0], dp[1]

    if l1 and entry_name == '合计':
        l1_total = None
        for e in entries:
            if e.get('isL1Total') and e.get('l1Name') == l1:
                l1_total = e
                break
        if not l1_total:
            return {'ok': False, 'value': '未匹配'}
        return _nz_resolve_metric(l1_total, col, metric, sum_col, ac_col, ac_val, fd)

    if entry_name == '总合计' or (not l1 and entry_name == '总合计'):
        return _nz_resolve_metric(total, col, metric, sum_col, ac_col, ac_val, fd)

    entry = None
    if l1:
        for e in entries:
            if e.get('isGroup') and e.get('l1Name') == l1 and e.get('name', '') == entry_name:
                entry = e
                break
        if not entry:
            for e in entries:
                if e.get('isL1Total') and e.get('l1Name') == l1:
                    entry = e
                    break
    else:
        for e in entries:
            if e.get('isGroup') and not e.get('l1Name') and e.get('name', '') == entry_name:
                entry = e
                break

    if not entry:
        return {'ok': False, 'value': '未匹配'}

    return _nz_resolve_metric(entry, col, metric, sum_col, ac_col, ac_val, fd)


def _nz_resolve_metric(entry, col, metric, sum_col, ac_col, ac_val, fd):
    if metric == '数量':
        if ac_col and ac_val:
            ac_data = entry.get('acData', {})
            tc = ac_data.get(ac_col, {})
            return {'ok': True, 'value': tc.get(ac_val, 0)}
        return {'ok': True, 'value': entry.get('count', 0)}
    if metric == '占比':
        pct = entry.get('pct', '0')
        try:
            pv = float(pct)
            return {'ok': True, 'value': _precise_round(pv), 'is_percent': True}
        except (ValueError, TypeError):
            return {'ok': False, 'value': '占比解析失败'}
    if metric == '求和':
        target_col = col or sum_col
        if not target_col:
            return {'ok': False, 'value': '未匹配(无求和列)'}
        s = entry.get('sum')
        if s is not None:
            try:
                return {'ok': True, 'value': float(s)}
            except (ValueError, TypeError):
                pass
        return {'ok': False, 'value': '未匹配(求和)'}
    if metric == '均值':
        target_col = col or sum_col
        if not target_col:
            return {'ok': False, 'value': '未匹配(无求和列)'}
        s = entry.get('sum')
        c = entry.get('count', 0)
        if s is not None and c > 0:
            try:
                return {'ok': True, 'value': round(float(s) / c, 2)}
            except (ValueError, TypeError, ZeroDivisionError):
                pass
        return {'ok': True, 'value': 0}
    return {'ok': False, 'value': '未知指标'}


# ========== 单元格引用解析 ==========

def _get_cell_value(ws_obj, col_str, row_str):
    try:
        col_idx = 0
        for ch in col_str:
            col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
        row_idx = int(row_str)
        if row_idx < 1 or col_idx < 1:
            return None
        c = ws_obj.cell(row=row_idx, column=col_idx)
        if c.value is None:
            return 0
        return c.value
    except (ValueError, TypeError):
        return None


def _col_letter_to_idx(col_str):
    """列字母转1-based索引（A=1, Z=26, AA=27）"""
    idx = 0
    for ch in col_str:
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx


def _resolve_cell_to_float(val, ws_obj, stats_data, visited=None):
    """将单元格值解析为float，支持纯数字/={{...}}表达式/混合公式文本。返回None表示无法解析。"""
    if val is None:
        return None
    val_str = str(val).strip()
    if not val_str:
        return None
    # = 开头的表达式（含单元格引用+{{}}+SUM/AVG）
    if val_str.startswith('='):
        expr = val_str[1:]
        expr, ok = _resolve_formulas_in_expr(expr, stats_data)
        if not ok:
            return None
        if visited is None:
            visited = set()
        expr = _resolve_funcs(expr, ws_obj, stats_data, visited)
        if 'NaN' in expr:
            return None
        expr = _resolve_cell_refs_in_expr(expr, ws_obj, stats_data, visited)
        if 'NaN' in expr:
            return None
        safe = re.sub(r'[^0-9+\-*/().eE\s]', '', expr)
        try:
            result = safe_eval_arithmetic(safe)
            if isinstance(result, (int, float)):
                return float(result)
        except (ValueError, SyntaxError, ZeroDivisionError):
            pass
        return None
    # 包含 {{...}} 公式（纯公式或混合文本）
    if _formula_pattern.search(val_str):
        resolved, ok = _resolve_formulas_in_text(val_str, stats_data)
        if not ok:
            return None
        try:
            return float(resolved)
        except (ValueError, TypeError):
            return None
    # 纯数字
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _resolve_cell_refs_in_expr(expr_str, ws_obj, stats_data, visited=None):
    if visited is None:
        visited = set()

    def replace_ref(m):
        col_s, row_s = m.group(1), m.group(2)
        ref_key = f"{col_s}{row_s}"
        if ref_key in visited:
            return 'NaN'
        visited.add(ref_key)
        val = _get_cell_value(ws_obj, col_s, row_s)
        if val is None:
            return 'NaN'
        num = _resolve_cell_to_float(val, ws_obj, stats_data, visited)
        if num is None:
            return 'NaN'
        return _fmt_num(num)

    return _cell_ref_pattern.sub(replace_ref, expr_str)


def _resolve_range_values(range_str, ws_obj, stats_data=None, visited=None):
    if visited is None:
        visited = set()
    values = []
    parts = [p.strip() for p in range_str.split(',')]
    range_re = re.compile(r'^([A-Z]{1,3})(\d{1,5}):([A-Z]{1,3})(\d{1,5})$')
    cell_re = re.compile(r'^([A-Z]{1,3})(\d{1,5})$')
    for part in parts:
        rm = range_re.match(part)
        if rm:
            c1 = _col_letter_to_idx(rm.group(1))
            r1 = int(rm.group(2))
            c2 = _col_letter_to_idx(rm.group(3))
            r2 = int(rm.group(4))
            min_r, max_r = min(r1, r2), max(r1, r2)
            min_c, max_c = min(c1, c2), max(c1, c2)
            for rr in range(min_r, max_r + 1):
                for cc in range(min_c, max_c + 1):
                    ref_key = f"{rr},{cc}"
                    if ref_key in visited:
                        continue
                    visited.add(ref_key)
                    cv = ws_obj.cell(row=rr, column=cc).value
                    num = _resolve_cell_to_float(cv, ws_obj, stats_data, visited)
                    if num is not None:
                        values.append(num)
        else:
            cm = cell_re.match(part)
            if cm:
                col_s, row_s = cm.group(1), cm.group(2)
                ref_key = f"{col_s}{row_s}"
                if ref_key not in visited:
                    visited.add(ref_key)
                    val = _get_cell_value(ws_obj, col_s, row_s)
                    num = _resolve_cell_to_float(val, ws_obj, stats_data, visited)
                    if num is not None:
                        values.append(num)
    return values


# ========== 数值精度处理 ==========

def _precise_sum(vals):
    """使用 math.fsum 消除浮点数累加精度误差，再 round 到合理精度"""
    result = math.fsum(vals)
    return _precise_round(result)


def _precise_round(value, max_decimals=10):
    """智能舍入：去除浮点尾数，保留有意义的位数（最多 max_decimals 位），最少保留 2 位。"""
    if not isinstance(value, (int, float)) or math.isnan(value) or math.isinf(value):
        return value
    if isinstance(value, int):
        return float(value)
    rounded = round(value, max_decimals)
    if rounded == int(rounded):
        return float(int(rounded))
    for d in range(2, max_decimals + 1):
        candidate = round(value, d)
        if round(candidate, max_decimals) == round(value, max_decimals):
            return candidate
    return rounded


def _fmt_num(value, decimal=None):
    """将数值格式化为字符串，控制小数位数。"""
    if not isinstance(value, (int, float)):
        return str(value)
    if math.isnan(value) or math.isinf(value):
        return str(value)
    if decimal is not None:
        rounded = round(value, decimal)
        if decimal == 0:
            return str(int(rounded))
        return str(rounded)
    v = _precise_round(value)
    if v == int(v):
        return str(int(v))
    return str(v)


# ========== 函数与表达式解析 ==========

def _resolve_funcs(expr_str, ws_obj, stats_data=None, visited=None):
    if visited is None:
        visited = set()

    def replace_func(m):
        fn = m.group(1).upper()
        args = m.group(2)
        vals = _resolve_range_values(args, ws_obj, stats_data, visited)
        if not vals:
            return 'NaN'
        if fn == 'SUM':
            return str(_precise_sum(vals))
        elif fn == 'AVG':
            return str(_precise_round(_precise_sum(vals) / len(vals)))
        return 'NaN'

    return _func_pattern.sub(replace_func, expr_str)


def _resolve_formulas_in_expr(expr, stats_data):
    all_ok = [True]
    def replace_fn(match):
        result = nz_resolve_formula_str(match.group(0), stats_data)
        if result['ok']:
            return _fmt_num(result['value'])
        all_ok[0] = False
        return 'NaN'
    resolved = _formula_pattern.sub(replace_fn, expr)
    return resolved, all_ok[0]


def _resolve_formulas_in_text(text, stats_data):
    any_fail = [False]
    def replace_fn(match):
        result = nz_resolve_formula_str(match.group(0), stats_data)
        if result['ok']:
            return _fmt_num(result['value'])
        any_fail[0] = True
        return match.group(0)
    resolved = _formula_pattern.sub(replace_fn, text)
    return resolved, not any_fail[0]


# ========== 核心：模板填充 ==========

def fill_template(template_bytes, stats_data, cell_edits, cell_formats):
    tmp_in = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_in.write(template_bytes)
    tmp_in.close()
    tmp_out = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_out.close()

    try:
        wb = load_workbook(tmp_in.name)
    except Exception as e:
        os.unlink(tmp_in.name)
        os.unlink(tmp_out.name)
        return {'ok': False, 'error': f'读取模板失败: {str(e)}'}

    _apply_cell_edits(wb, cell_edits)
    _apply_cell_formats(wb, cell_formats)

    fill_count = 0
    fail_count = 0
    percent_value_cells = set()

    for ws_idx, ws in enumerate(wb.worksheets):
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_str = str(cell.value).strip()

                has_formula = _formula_pattern.search(cell_str) or cell_str.startswith('=') or re.search(r'=(?:SUM|AVG)\s*\(', cell_str, re.IGNORECASE)
                if not has_formula:
                    continue

                if cell_str.startswith('='):
                    fill_count, fail_count = _fill_expression_cell(
                        ws, ws_idx, cell, cell_str, stats_data, cell_formats,
                        percent_value_cells, fill_count, fail_count)
                elif re.match(r'^\{\{(.+?)\}\}$', cell_str):
                    fill_count, fail_count = _fill_pure_formula_cell(
                        ws_idx, cell, cell_str, stats_data, percent_value_cells,
                        fill_count, fail_count)
                else:
                    fill_count, fail_count = _fill_mixed_text_cell(
                        ws, ws_idx, cell, cell_str, stats_data, cell_formats,
                        fill_count, fail_count)

    _apply_number_formats(wb, cell_formats, percent_value_cells)

    try:
        wb.save(tmp_out.name)
        wb.close()
    except Exception as e:
        wb.close()
        os.unlink(tmp_in.name)
        os.unlink(tmp_out.name)
        return {'ok': False, 'error': f'保存失败: {str(e)}'}

    return {
        'ok': True,
        'tmp_in': tmp_in.name,
        'tmp_out': tmp_out.name,
        'fill_count': fill_count,
        'fail_count': fail_count
    }


def _apply_cell_edits(wb, cell_edits):
    """应用前端传来的单元格编辑（直接覆盖值）。"""
    for edit in cell_edits:
        si = edit.get('sheet', 0)
        r = edit.get('row', 0)
        c = edit.get('col', 0)
        val = edit.get('value', '')
        if si < 0 or si >= len(wb.sheetnames):
            continue
        ws = wb.worksheets[si]
        cell = ws.cell(row=r + 1, column=c + 1)
        try:
            num_val = float(val) if val and '.' in val else (int(val) if val and val.lstrip('-').isdigit() else None)
        except (ValueError, TypeError, AttributeError):
            num_val = None
        if num_val is not None:
            cell.value = num_val
        else:
            cell.value = val


def _apply_cell_formats(wb, cell_formats):
    """应用前端传来的单元格字体/对齐格式。"""
    for fmt_item in cell_formats:
        si = fmt_item.get('sheet', 0)
        r = fmt_item.get('row', 0)
        c = fmt_item.get('col', 0)
        fmt = fmt_item.get('fmt', {})
        if si < 0 or si >= len(wb.sheetnames):
            continue
        ws = wb.worksheets[si]
        cell = ws.cell(row=r + 1, column=c + 1)
        if fmt.get('bold') or fmt.get('italic') or fmt.get('fontSize') or fmt.get('fontName'):
            cell.font = cell.font.copy(
                bold=fmt.get('bold', cell.font.bold),
                italic=fmt.get('italic', cell.font.italic),
                size=fmt.get('fontSize', cell.font.size),
                name=fmt.get('fontName', cell.font.name)
            )
        if fmt.get('align'):
            align_map = {'left': 'left', 'center': 'center', 'right': 'right'}
            cell.alignment = cell.alignment.copy(horizontal=align_map.get(fmt['align'], cell.alignment.horizontal))


def _fill_expression_cell(ws, ws_idx, cell, cell_str, stats_data, cell_formats,
                           percent_value_cells, fill_count, fail_count):
    """填充 = 开头的表达式单元格。"""
    expr = cell_str[1:]
    expr, all_ok = _resolve_formulas_in_expr(expr, stats_data)
    if not all_ok:
        cell.value = '公式解析失败'
        return fill_count, fail_count + 1

    visited = set()
    self_col = get_column_letter(cell.column)
    self_row = str(cell.row)
    visited.add(f"{self_col}{self_row}")
    expr = _resolve_funcs(expr, ws, stats_data, visited)
    if 'NaN' in expr:
        cell.value = '函数解析失败'
        return fill_count, fail_count + 1

    expr = _resolve_cell_refs_in_expr(expr, ws, stats_data, visited)
    if 'NaN' in expr:
        cell.value = '引用解析失败'
        return fill_count, fail_count + 1

    safe_expr = re.sub(r'[^0-9+\-*/().eE\s]', '', expr)
    try:
        num_result = safe_eval_arithmetic(safe_expr)
        if isinstance(num_result, (int, float)) and str(num_result) != 'nan':
            if isinstance(num_result, float) and not num_result == int(num_result):
                num_result = _precise_round(num_result)
            cell.value = num_result
            return fill_count + 1, fail_count
        else:
            cell.value = '计算错误'
            return fill_count, fail_count + 1
    except (ValueError, SyntaxError, ZeroDivisionError):
        cell.value = '表达式错误'
        return fill_count, fail_count + 1


def _fill_pure_formula_cell(ws_idx, cell, cell_str, stats_data, percent_value_cells,
                             fill_count, fail_count):
    """填充纯 {{...}} 公式单元格。"""
    result = nz_resolve_formula_str(cell_str, stats_data)
    if result['ok']:
        val = result['value']
        if isinstance(val, float) and not val == int(val):
            val = _precise_round(val)
        cell.value = val
        if result.get('is_percent'):
            percent_value_cells.add((ws_idx, cell.row, cell.column))
        return fill_count + 1, fail_count
    else:
        cell.value = str(result['value'])
        return fill_count, fail_count + 1


def _fill_mixed_text_cell(ws, ws_idx, cell, cell_str, stats_data, cell_formats,
                           fill_count, fail_count):
    """填充混合文本单元格（含 {{...}} 和/或 =SUM/=AVG）。"""
    final_val, ok = _resolve_formulas_in_text(cell_str, stats_data)

    if re.search(r'=(SUM|AVG)\s*\(', final_val, re.IGNORECASE) or re.search(r'=[A-Z]+\d+', final_val):
        _cell_key = (ws_idx, cell.row - 1, cell.column - 1)
        _cell_fmt = {}
        for fmt_item in cell_formats:
            _fk = (fmt_item.get('sheet', 0), fmt_item.get('row', 0), fmt_item.get('col', 0))
            if _fk == _cell_key:
                _cell_fmt = fmt_item.get('fmt', {})
                break
        _cell_decimal = _cell_fmt.get('decimal', None)

        visited = set()
        self_col = get_column_letter(cell.column)
        self_row = str(cell.row)
        visited.add(f"{self_col}{self_row}")
        final_val = _resolve_funcs(final_val, ws, stats_data, visited)
        final_val = _resolve_cell_refs_in_expr(final_val, ws, stats_data, visited)

        def _eval_formula_part(m):
            expr_str = m.group(0)[1:]
            try:
                safe = re.sub(r'[^0-9+\-*/().eE\s]', '', expr_str)
                result = safe_eval_arithmetic(safe)
                if isinstance(result, float) and not result == int(result):
                    d = _cell_decimal if _cell_decimal is not None else 2
                    rounded = round(result, d)
                    return str(int(rounded)) if d == 0 else str(rounded)
                return str(int(result)) if isinstance(result, float) else str(result)
            except (ValueError, SyntaxError, ZeroDivisionError, TypeError):
                return m.group(0)
        final_val = re.sub(r'=(?:SUM|AVG)\s*\([^)]*\)', _eval_formula_part, final_val, flags=re.IGNORECASE)

    cell.value = final_val
    if not ok:
        return fill_count, fail_count + 1
    return fill_count + 1, fail_count


def _apply_number_formats(wb, cell_formats, percent_value_cells):
    """应用小数位数/百分比数字格式 + 自动适配模板原生百分比。"""
    fmt_lookup = {}
    for fmt_item in cell_formats:
        key = (fmt_item.get('sheet', 0), fmt_item.get('row', 0), fmt_item.get('col', 0))
        fmt_lookup[key] = fmt_item.get('fmt', {})

    for ws_idx, ws in enumerate(wb.worksheets):
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                key = (ws_idx, cell.row - 1, cell.column - 1)
                fmt = fmt_lookup.get(key, {})
                decimal = fmt.get('decimal', None)
                isPercent = fmt.get('percent', False)
                if decimal is None and not isPercent:
                    continue
                try:
                    v = float(cell.value)
                except (ValueError, TypeError):
                    continue
                if isPercent:
                    d = decimal if decimal is not None else 2
                    factor = 10 ** d
                    pct_rounded = round(v * 100 * factor) / factor
                    ratio = pct_rounded / 100
                    cell.value = int(ratio) if d == 0 else ratio
                    cell.number_format = '0%' if d == 0 else f'0.{"0" * d}%'
                elif decimal is not None:
                    factor = 10 ** decimal
                    rounded = round(v * factor) / factor
                    cell.value = int(rounded) if decimal == 0 else rounded
                    cell.number_format = '0' if decimal == 0 else f'0.{"0" * decimal}'

    # 自动适配模板原生百分比格式
    for (ws_i, r, c) in percent_value_cells:
        fmt_key = (ws_i, r - 1, c - 1)
        if fmt_key in fmt_lookup and fmt_lookup[fmt_key].get('percent'):
            continue
        ws = wb.worksheets[ws_i]
        cell = ws.cell(row=r, column=c)
        fmt = cell.number_format or ''
        if not ('%' in fmt and '"%"' not in fmt):
            try:
                float(cell.value)
                cell.number_format = '0.00%'
            except (ValueError, TypeError):
                pass
