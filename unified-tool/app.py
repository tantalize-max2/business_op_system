# -*- coding: utf-8 -*-
"""
商机数据综合分析工具 - Flask 后端
功能：分局拆分（基于过滤后数据）、格式保持、文件下载
"""

import os
import json
import re
import random
import shutil
import zipfile
import tempfile
import base64
import http.client
import ssl
import time
import uuid
import threading
import urllib.parse
from datetime import datetime
from urllib.parse import urlparse
from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

app = Flask(__name__, static_folder='static', static_url_path='')
CORS(app)

# ============ 数据存储 ============
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
UPLOAD_DIR = os.path.join(DATA_DIR, 'uploads')
OUTPUT_DIR = os.path.join(DATA_DIR, 'output')
CONFIGS_DIR = os.path.join(DATA_DIR, 'configs')
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(CONFIGS_DIR, exist_ok=True)

MAPPING_FILE = os.path.join(DATA_DIR, 'bureau_mapping.json')
SHEETS_FILE = os.path.join(DATA_DIR, 'kdocs_sheets.json')
TEMPLATES_DIR = os.path.join(DATA_DIR, 'bureau_templates')
NZ_TEMPLATES_DIR = os.path.join(DATA_DIR, 'nz_templates')
os.makedirs(TEMPLATES_DIR, exist_ok=True)
os.makedirs(NZ_TEMPLATES_DIR, exist_ok=True)

# ============ 默认分局人员映射 ============
DEFAULT_MAPPING = {
    "工业能源政企分局": ["朱晨静", "肖智宇", "杜云帆", "屈容", "陈谦", "唐璐", "杜秋宇"],
    "国有平台政企分局": ["付登会", "蓝旭辉", "田小兰", "颜小琳", "易琴", "杨文", "王越"],
    "健康医疗政企分局": ["潘昱忻", "黎明", "先有为", "王虹丹"],
    "金融证券政企分局": ["贺贤珍", "曹岚", "梁曦宇", "杨彭萱", "张静", "郑文", "刘浩林", "杨真", "杨玮萍"],
    "软件科研政企分局": ["雷世豪", "陈燕", "杨璨宇", "李春江", "潘邓浩", "蒋尧莉", "李永婷", "蒲竑佚", "周建龙", "杨楚琪"],
    "新经济政企分局": ["曾理", "韩思萌", "罗维", "罗霞", "杨佩东", "刘文博", "戚新鹏"],
    "政法应急政企分局": ["王万辞", "徐杨", "袁进", "朱林", "袁新博", "张皓秋", "林佳俊", "黄天玉"],
    "政务政企分局": ["薛笑枫", "廖晓东", "钱宇煊", "吴倩", "杜成思", "黄振宇"],
    "高新孵化园智改数转服务局": ["董小凤", "邱国锋", "龙俊儒"],
    "高新天府生命科技园智改数转服务局": ["杨佳", "蒋建国", "王斯祺"],
    "高新天府软件园智改数转服务局": ["谢思宇", "黄微", "李选玉", "朱勇", "谢勇", "何琼", "高欢", "周莉", "李艺"],
    "金融城商客分局": ["姚尧", "王淑惠", "陈伟智", "曾宇嘉", "罗中伟", "张成铭", "刘荣"],
    "新川商客分局": ["靳扬", "姜春阳", "郑黎霞", "杨晋"],
    "天府新谷商客分局": ["李思锐", "彭倩", "黄燕", "刘鹏洋", "刘星月"],
    "新会展商客分局": ["何艳", "樊志林", "周滨", "蒋稚薇"],
    "天府国际商客分局": ["叶江", "蒋天佑", "杨茜", "李巧巧", "王辰雨", "廖华", "曾明全", "巫婷婷"],
    "环球商客分局": ["曾明", "许可", "钟小燕", "肖福洋", "冯麟霞", "王琴丽"],
    "大源商客分局": ["邱浩锋", "冯兰越", "冯特峰", "裴嘉轩", "何亚琪"],
    "肖芳商客分局": ["任登科", "贾小东", "梁润", "陈雪"],
    "府城商客分局": ["陈磊", "李若玉", "张小龙", "孙雯"],
    "连锁商客分局": ["杨凤翥", "肖帆", "赵娇", "高毛茅", "温有军"],
    "西信商客分局": ["杨力", "王宇", "任少杰", "周雨晴", "刘祖源", "胡文瀚", "赵川川"],
    "东苑商客分局": ["雷蕾", "吴文宪", "孙艺丹", "张宇魁", "聂海林", "陈健明"],
    "校园分局": ["薛程月", "李经霜", "阳文婷", "欧阳晨"]
}


def load_mapping():
    if os.path.exists(MAPPING_FILE):
        try:
            with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return DEFAULT_MAPPING.copy()


def save_mapping(mapping):
    with open(MAPPING_FILE, 'w', encoding='utf-8') as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)


def clean_name(name):
    if not isinstance(name, str):
        return ""
    name = re.sub(r'\([^)]*\)', '', name)
    name = re.sub(r'\（[^）]*\）', '', name)
    # 清理常见分隔符（逗号、顿号、分号等），保留首个名字
    # 与前端预处理逻辑保持一致：按分隔符拆分后取第一个非空值
    parts = re.split(r'[,，、;；\s]+', name)
    parts = [p.strip() for p in parts if p.strip()]
    return parts[0] if parts else name.strip()


def copy_row_with_format(source_sheet, target_sheet, source_row, target_row, max_col):
    for col in range(1, max_col + 1):
        source_cell = source_sheet.cell(row=source_row, column=col)
        target_cell = target_sheet.cell(row=target_row, column=col)
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell.font = source_cell.font.copy()
            target_cell.border = source_cell.border.copy()
            target_cell.fill = source_cell.fill.copy()
            target_cell.number_format = source_cell.number_format
            target_cell.protection = source_cell.protection.copy()
            target_cell.alignment = source_cell.alignment.copy()
        col_letter = get_column_letter(col)
        target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width


def copy_sheet_with_format(source_sheet, target_sheet, row_indices):
    max_col = source_sheet.max_column
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width
    target_row = 1
    for source_row in row_indices:
        copy_row_with_format(source_sheet, target_sheet, source_row, target_row, max_col)
        if source_sheet.row_dimensions[source_row].height:
            target_sheet.row_dimensions[target_row].height = source_sheet.row_dimensions[source_row].height
        target_row += 1
    target_sheet.sheet_format = source_sheet.sheet_format
    target_sheet.page_setup = source_sheet.page_setup
    target_sheet.page_margins = source_sheet.page_margins


# ============ API 路由 ============

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')


@app.route('/api/mapping', methods=['GET'])
def get_mapping():
    data = load_mapping()
    # 使用 json.dumps 而非 jsonify，保留键的插入顺序（不排序）
    return app.response_class(
        response=json.dumps(data, ensure_ascii=False),
        status=200,
        mimetype='application/json'
    )


@app.route('/api/mapping', methods=['POST'])
def save_mapping_api():
    mapping = request.json
    if not mapping or not isinstance(mapping, dict):
        return jsonify({'error': '无效的映射数据'}), 400
    save_mapping(mapping)
    return jsonify({'message': '保存成功'})


@app.route('/api/reset-mapping', methods=['POST'])
def reset_mapping():
    save_mapping(DEFAULT_MAPPING.copy())
    return jsonify({'message': '已重置为默认映射'})


# ============ 分局模板 API ============

def _template_path(name):
    safe_name = re.sub(r'[^\w\u4e00-\u9fff\-\.]', '_', name)
    return os.path.join(TEMPLATES_DIR, f"{safe_name}.json")


@app.route('/api/bureau-templates', methods=['GET'])
def list_bureau_templates():
    templates = []
    for fname in os.listdir(TEMPLATES_DIR):
        if not fname.endswith('.json'):
            continue
        fpath = os.path.join(TEMPLATES_DIR, fname)
        try:
            with open(fpath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            templates.append({
                'name': data.get('name', fname[:-5]),
                'bureauCount': len(data.get('mapping', {})),
                'savedAt': data.get('savedAt', 0)
            })
        except:
            pass
    templates.sort(key=lambda t: t.get('savedAt', 0), reverse=True)
    return jsonify(templates)


@app.route('/api/bureau-templates', methods=['POST'])
def save_bureau_template():
    data = request.json or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': '模板名称不能为空'}), 400
    mapping = data.get('mapping')
    if not mapping or not isinstance(mapping, dict):
        return jsonify({'error': '映射数据不能为空'}), 400
    template_data = {
        'name': name,
        'mapping': mapping,
        'savedAt': data.get('savedAt', datetime.now().timestamp() * 1000)
    }
    fpath = _template_path(name)
    with open(fpath, 'w', encoding='utf-8') as f:
        json.dump(template_data, f, ensure_ascii=False, indent=2)
    return jsonify({'message': '模板已保存', 'name': name})


@app.route('/api/bureau-templates/<path:name>', methods=['GET'])
def get_bureau_template(name):
    fpath = _template_path(name)
    if not os.path.exists(fpath):
        return jsonify({'error': '模板不存在'}), 404
    with open(fpath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return app.response_class(
        response=json.dumps(data, ensure_ascii=False),
        status=200,
        mimetype='application/json'
    )


@app.route('/api/bureau-templates/<path:name>', methods=['DELETE'])
def delete_bureau_template(name):
    fpath = _template_path(name)
    if not os.path.exists(fpath):
        return jsonify({'error': '模板不存在'}), 404
    os.remove(fpath)
    return jsonify({'message': '模板已删除'})


# ============ 配置管理 API ============

def _config_path(name):
    """安全获取配置文件路径，防止路径穿越"""
    safe_name = re.sub(r'[^\w\u4e00-\u9fff\-\.]', '_', name)
    return os.path.join(CONFIGS_DIR, f"{safe_name}.json")


@app.route('/api/configs', methods=['GET'])
def list_configs():
    """列出所有已保存配置"""
    configs = []
    for fname in os.listdir(CONFIGS_DIR):
        if not fname.endswith('.json'):
            continue
        fpath = os.path.join(CONFIGS_DIR, fname)
        try:
            with open(fpath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            cfg = data.get('cfg', {})
            mapping_data = cfg.get('mappingData')
            configs.append({
                'name': data.get('name', fname[:-5]),
                'sig': data.get('sig', ''),
                'savedAt': data.get('savedAt', 0),
                'fileNames': [fd.get('name', '') for fd in cfg.get('files', [])],
                'hasMapping': mapping_data is not None and isinstance(mapping_data, dict) and len(mapping_data) > 0,
                'mappingCount': len(mapping_data) if isinstance(mapping_data, dict) else 0
            })
        except:
            pass
    configs.sort(key=lambda c: c.get('savedAt', 0), reverse=True)
    return app.response_class(
        response=json.dumps(configs, ensure_ascii=False),
        status=200,
        mimetype='application/json'
    )


@app.route('/api/configs', methods=['POST'])
def save_config():
    """保存配置"""
    data = request.json or {}
    name = data.get('name', '').strip()
    cfg = data.get('cfg')
    sig = data.get('sig', '')
    if not name:
        return jsonify({'error': '配置名称不能为空'}), 400
    if not cfg:
        return jsonify({'error': '配置数据不能为空'}), 400
    config_data = {
        'name': name,
        'sig': sig,
        'cfg': cfg,
        'savedAt': data.get('savedAt', datetime.now().timestamp() * 1000)
    }
    fpath = _config_path(name)
    # 限制最多20个配置
    existing = [f for f in os.listdir(CONFIGS_DIR) if f.endswith('.json')]
    if len(existing) >= 20 and not os.path.exists(fpath):
        # 删除最旧的
        all_configs = []
        for ef in existing:
            ep = os.path.join(CONFIGS_DIR, ef)
            try:
                with open(ep, 'r', encoding='utf-8') as f:
                    d = json.load(f)
                all_configs.append((ef, d.get('savedAt', 0)))
            except:
                all_configs.append((ef, 0))
        all_configs.sort(key=lambda x: x[1])
        os.remove(os.path.join(CONFIGS_DIR, all_configs[0][0]))
    with open(fpath, 'w', encoding='utf-8') as f:
        json.dump(config_data, f, ensure_ascii=False)
    return jsonify({'message': '配置已保存', 'name': name})


@app.route('/api/configs/<path:name>', methods=['GET'])
def get_config(name):
    """获取指定配置"""
    fpath = _config_path(name)
    if not os.path.exists(fpath):
        return jsonify({'error': '配置不存在'}), 404
    with open(fpath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return app.response_class(
        response=json.dumps(data, ensure_ascii=False),
        status=200,
        mimetype='application/json'
    )


@app.route('/api/configs/<path:name>', methods=['DELETE'])
def delete_config(name):
    """删除指定配置"""
    fpath = _config_path(name)
    if not os.path.exists(fpath):
        return jsonify({'error': '配置不存在'}), 404
    os.remove(fpath)
    return jsonify({'message': '配置已删除'})


@app.route('/api/split-filtered', methods=['POST'])
def split_filtered():
    """对前端一级过滤后的数据进行分局拆分

    接收: {
      fileName: str,               # 原始文件名
      fileDataBase64: str,         # 原始Excel文件的base64编码
      filteredRowIndices: [int],   # 一级过滤后的行索引(0-based, 不含表头)
      mapping: dict                # 分局映射(可选, 默认用保存的)
    }
    """
    data = request.json or {}
    file_data_b64 = data.get('fileDataBase64', '')
    filtered_indices = data.get('filteredRowIndices', [])
    mapping = data.get('mapping') or load_mapping()
    split_column = data.get('splitColumn', '')

    if not file_data_b64:
        return jsonify({'error': '缺少文件数据'}), 400

    # 解码 base64 -> 临时文件
    import base64
    try:
        raw_bytes = base64.b64decode(file_data_b64)
    except:
        return jsonify({'error': '文件数据解码失败'}), 400

    tmp_in = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_in.write(raw_bytes)
    tmp_in.close()

    # 清理旧输出
    for f in os.listdir(OUTPUT_DIR):
        fp = os.path.join(OUTPUT_DIR, f)
        if os.path.isfile(fp):
            os.remove(fp)
        elif os.path.isdir(fp):
            shutil.rmtree(fp)

    current_date = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_folder = os.path.join(OUTPUT_DIR, f"分局拆分结果_{current_date}")
    os.makedirs(output_folder, exist_ok=True)

    try:
        source_wb = load_workbook(tmp_in.name)
        source_sheet = source_wb.active
    except Exception as e:
        os.unlink(tmp_in.name)
        return jsonify({'error': f'读取文件出错: {str(e)}'}), 500

    try:
        df = pd.read_excel(tmp_in.name, sheet_name=source_sheet.title, dtype=str)
    except Exception as e:
        source_wb.close()
        os.unlink(tmp_in.name)
        return jsonify({'error': f'读取数据出错: {str(e)}'}), 500

    # 只在过滤后的行中进行匹配
    filtered_set = set(filtered_indices) if filtered_indices else set(range(len(df)))

    if not split_column:
        return jsonify({'error': '未指定拆分列'}), 400
    if split_column not in df.columns:
        return jsonify({'error': f'拆分列 "{split_column}" 不存在于数据中'}), 400

    bureau_rows = {bureau: [] for bureau in mapping.keys()}
    unmatched_rows = []
    header_row = 1
    matched_count = 0
    unmatched_count = 0
    unmatched_managers = set()

    for index in filtered_set:
        if index < 0 or index >= len(df):
            continue
        excel_row = index + 2  # 转为Excel行号(1-based, +1表头)

        manager_name_raw = df.iloc[index][split_column]
        manager_name_clean = clean_name(manager_name_raw)
        matched = False
        for bureau, managers in mapping.items():
            if manager_name_clean in managers:
                bureau_rows[bureau].append(excel_row)
                matched_count += 1
                matched = True
                break

        if not matched:
            unmatched_rows.append(excel_row)
            unmatched_count += 1
            if manager_name_clean:
                unmatched_managers.add(manager_name_clean)

    # 为每个分局生成文件
    generated_files = []
    for bureau_name, row_indices in bureau_rows.items():
        if row_indices:
            target_wb = Workbook()
            target_sheet = target_wb.active
            all_rows = [header_row] + row_indices
            copy_sheet_with_format(source_sheet, target_sheet, all_rows)
            safe_name = re.sub(r'[\/\\:*?"<>|]', '_', bureau_name)
            output_file = os.path.join(output_folder, f"{safe_name}_{current_date}.xlsx")
            target_wb.save(output_file)
            generated_files.append({
                'bureau': bureau_name,
                'rows': len(row_indices),
                'filename': f"{safe_name}_{current_date}.xlsx"
            })

    # 汇总文件
    summary_file = os.path.join(output_folder, f"行业商机数据汇总_{current_date}.xlsx")
    summary_wb = Workbook()
    if 'Sheet' in summary_wb.sheetnames:
        del summary_wb['Sheet']
    for bureau_name, row_indices in bureau_rows.items():
        sheet_name = bureau_name[:31]
        target_sheet = summary_wb.create_sheet(title=sheet_name)
        if row_indices:
            all_rows = [header_row] + row_indices
            copy_sheet_with_format(source_sheet, target_sheet, all_rows)
    summary_wb.save(summary_file)
    generated_files.append({
        'bureau': '- 汇总文件 -',
        'rows': matched_count,
        'filename': f"行业商机数据汇总_{current_date}.xlsx"
    })

    # 行业/商业分类汇总
    INDUSTRY_BUREAUS = [
        "政法应急政企分局", "国有平台政企分局", "金融证券政企分局", "工业能源政企分局",
        "政务政企分局", "高新天府软件园智改数转服务局", "软件科研政企分局",
        "高新孵化园智改数转服务局", "新经济政企分局", "健康医疗政企分局",
        "高新天府生命科技园智改数转服务局"
    ]
    COMMERCIAL_BUREAUS = [
        "校园分局", "新川商客分局", "金融城商客分局", "肖芳商客分局",
        "东苑商客分局", "天府国际商客分局", "大源商客分局", "新会展商客分局",
        "连锁商客分局", "环球商客分局", "天府新谷商客分局", "西信商客分局", "府城商客分局"
    ]

    def make_category_summary(category_name, category_bureaus):
        """生成行业/商业汇总文件，每个分局一个sheet"""
        cat_wb = Workbook()
        if 'Sheet' in cat_wb.sheetnames:
            del cat_wb['Sheet']
        cat_rows = 0
        for bureau_name in category_bureaus:
            row_indices = bureau_rows.get(bureau_name, [])
            if row_indices:
                sheet_name = bureau_name[:31]
                target_sheet = cat_wb.create_sheet(title=sheet_name)
                all_rows = [header_row] + row_indices
                copy_sheet_with_format(source_sheet, target_sheet, all_rows)
                cat_rows += len(row_indices)
        if cat_rows:
            safe_cat = re.sub(r'[\/\\:*?"<>|]', '_', category_name)
            cat_file = os.path.join(output_folder, f"{safe_cat}_{current_date}.xlsx")
            cat_wb.save(cat_file)
            generated_files.append({
                'bureau': f'- {category_name} -',
                'rows': cat_rows,
                'filename': f"{safe_cat}_{current_date}.xlsx"
            })

    make_category_summary("行业数据汇总", INDUSTRY_BUREAUS)
    make_category_summary("商业数据汇总", COMMERCIAL_BUREAUS)

    # 未匹配名单
    if unmatched_rows:
        unmatched_wb = Workbook()
        unmatched_sheet = unmatched_wb.active
        all_rows = [header_row] + unmatched_rows
        copy_sheet_with_format(source_sheet, unmatched_sheet, all_rows)
        unmatched_file_path = os.path.join(output_folder, f"未匹配名单_{current_date}.xlsx")
        unmatched_wb.save(unmatched_file_path)
        generated_files.append({
            'bureau': '- 未匹配名单 -',
            'rows': unmatched_count,
            'filename': f"未匹配名单_{current_date}.xlsx"
        })

    # ZIP打包
    zip_name = f"分局拆分结果_{current_date}.zip"
    zip_path = os.path.join(OUTPUT_DIR, zip_name)
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for f in os.listdir(output_folder):
            zf.write(os.path.join(output_folder, f), f)

    source_wb.close()
    os.unlink(tmp_in.name)

    return jsonify({
        'message': '拆分完成',
        'matched': matched_count,
        'unmatched': unmatched_count,
        'totalFiltered': len(filtered_set),
        'files': generated_files,
        'unmatched_managers': sorted(unmatched_managers),
        'zip': zip_name,
        'output_folder': f"分局拆分结果_{current_date}"
    })


@app.route('/api/download/<path:filename>')
def download_file(filename):
    return send_file(os.path.join(OUTPUT_DIR, filename), as_attachment=True)


@app.route('/api/download-folder/<path:folder>')
def download_folder(folder):
    folder_path = os.path.join(OUTPUT_DIR, folder)
    if not os.path.isdir(folder_path):
        return jsonify({'error': '文件夹不存在'}), 404
    zip_name = f"{folder}.zip"
    zip_path = os.path.join(OUTPUT_DIR, zip_name)
    if not os.path.exists(zip_path):
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for f in os.listdir(folder_path):
                zf.write(os.path.join(folder_path, f), f)
    return send_file(zip_path, as_attachment=True)


# ============ 在线表格管理 API ============

KDOCS_CATS_FILE = os.path.join(DATA_DIR, 'kdocs_categories.json')

def _load_kdocs_sheets():
    """加载在线表格配置列表"""
    if os.path.exists(SHEETS_FILE):
        try:
            with open(SHEETS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return []

def _save_kdocs_sheets(sheets):
    """保存在线表格配置列表"""
    with open(SHEETS_FILE, 'w', encoding='utf-8') as f:
        json.dump(sheets, f, ensure_ascii=False, indent=2)

def _load_kdocs_cats():
    """加载分类列表"""
    if os.path.exists(KDOCS_CATS_FILE):
        try:
            with open(KDOCS_CATS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return [{'id': 'default', 'name': '默认', 'color': '#0d9488'}]

def _save_kdocs_cats(cats):
    """保存分类列表"""
    with open(KDOCS_CATS_FILE, 'w', encoding='utf-8') as f:
        json.dump(cats, f, ensure_ascii=False, indent=2)


@app.route('/api/kdocs-browse', methods=['POST'])
def browse_local_fs():
    """浏览本地文件系统，返回指定路径下的子目录和xlsx文件"""
    data = request.json or {}
    path = data.get('path', '').strip()

    # 特殊路径 "__drives__" 表示列出所有盘符
    if path == '__drives__':
        drives = []
        for letter in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
            d = f'{letter}:\\'
            if os.path.exists(d):
                drives.append({'name': f'{letter}:', 'path': d, 'is_drive': True})
        return jsonify({
            'current': '__drives__',
            'current_display': '此电脑',
            'parent': '',
            'dirs': drives,
            'files': [],
            'is_drives': True
        })

    if not path:
        path = os.path.expanduser('~')
    if not os.path.exists(path):
        return jsonify({'error': f'路径不存在: {path}'}), 400

    dirs = []
    files = []
    try:
        for item in os.listdir(path):
            full = os.path.join(path, item)
            if os.path.isdir(full):
                dirs.append({'name': item, 'path': full})
            elif item.endswith(('.xlsx', '.xls')) and not item.startswith('~$'):
                files.append({'name': item, 'path': full, 'size': os.path.getsize(full)})
    except PermissionError:
        return jsonify({'error': '无权限访问该路径'}), 400

    # 判断上级目录：盘符根目录的上级回到盘符列表
    parent = ''
    if path and len(path) <= 3 and path.endswith(':\\'):
        # 盘符根目录，上级回到盘符列表
        parent = '__drives__'
    elif path:
        parent = os.path.dirname(path)
        if parent == path:
            parent = '__drives__'

    return jsonify({
        'current': path,
        'parent': parent,
        'dirs': sorted(dirs, key=lambda x: x['name'].lower()),
        'files': sorted(files, key=lambda x: x['name'].lower()),
        'is_dir': os.path.isdir(path)
    })


@app.route('/api/kdocs-categories', methods=['GET'])
def list_kdocs_cats():
    """列出所有分类"""
    cats = _load_kdocs_cats()
    sheets = _load_kdocs_sheets()
    # 附加每个分类下有多少个表格
    for c in cats:
        c['count'] = sum(1 for s in sheets if s.get('category') == c['id'])
    return jsonify(cats)


@app.route('/api/kdocs-categories', methods=['POST'])
def add_kdocs_cat():
    """添加分类"""
    data = request.json or {}
    name = data.get('name', '').strip()
    color = data.get('color', '#0d9488').strip()
    if not name:
        return jsonify({'error': '分类名不能为空'}), 400
    cats = _load_kdocs_cats()
    if any(c['name'] == name for c in cats):
        return jsonify({'error': '分类名已存在'}), 400
    cat = {'id': str(uuid.uuid4())[:8], 'name': name, 'color': color}
    cats.append(cat)
    _save_kdocs_cats(cats)
    return jsonify({'message': '添加成功', 'category': cat})


@app.route('/api/kdocs-categories/<cid>', methods=['DELETE'])
def delete_kdocs_cat(cid):
    """删除分类"""
    cats = _load_kdocs_cats()
    # 不允许删除默认分类
    cats = [c for c in cats if not (c['id'] == cid and c['id'] == 'default')]
    _save_kdocs_cats(cats)
    # 将属于该分类的表格移至默认分类
    sheets = _load_kdocs_sheets()
    for s in sheets:
        if s.get('category') == cid:
            s['category'] = 'default'
    _save_kdocs_sheets(sheets)
    return jsonify({'message': '已删除'})


@app.route('/api/kdocs-sheets', methods=['GET'])
def list_kdocs_sheets():
    """列出所有在线表格配置"""
    sheets = _load_kdocs_sheets()
    cat_id = request.args.get('category', '')
    if cat_id:
        sheets = [s for s in sheets if s.get('category', 'default') == cat_id]
    return jsonify(sheets)


@app.route('/api/kdocs-sheets', methods=['POST'])
def add_kdocs_sheet():
    """添加在线表格配置"""
    data = request.json or {}
    name = data.get('name', '').strip()
    url = data.get('url', '').strip()
    api_token = data.get('api_token', '').strip()
    webhook_url = data.get('webhook_url', '').strip()
    excel_path = data.get('excel_path', '').strip()
    batch_size = data.get('batch_size', 3)
    category = data.get('category', 'default')

    if not name or not url:
        return jsonify({'error': '名称和URL不能为空'}), 400

    sheets = _load_kdocs_sheets()

    sheet = {
        'id': str(uuid.uuid4())[:8],
        'name': name,
        'url': url,
        'api_token': api_token,
        'webhook_url': webhook_url,
        'excel_path': excel_path,
        'batch_size': batch_size,
        'category': category,
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    sheets.append(sheet)
    _save_kdocs_sheets(sheets)
    return jsonify({'message': '添加成功', 'sheet': sheet})


@app.route('/api/kdocs-sheets/<sid>', methods=['PUT'])
def update_kdocs_sheet(sid):
    """更新在线表格配置"""
    data = request.json or {}
    sheets = _load_kdocs_sheets()
    for s in sheets:
        if s['id'] == sid:
            for k in ['name', 'url', 'api_token', 'webhook_url', 'excel_path', 'batch_size', 'category']:
                if k in data:
                    s[k] = data[k]
            s['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            _save_kdocs_sheets(sheets)
            return jsonify({'message': '更新成功', 'sheet': s})
    return jsonify({'error': '未找到该配置'}), 404


@app.route('/api/kdocs-sheets/<sid>', methods=['DELETE'])
def delete_kdocs_sheet(sid):
    """删除在线表格配置"""
    sheets = _load_kdocs_sheets()
    sheets = [s for s in sheets if s['id'] != sid]
    _save_kdocs_sheets(sheets)
    return jsonify({'message': '删除成功'})


@app.route('/api/kdocs-push', methods=['POST'])
def push_to_kdocs():
    """将本地Excel数据推送到金山文档在线表格"""
    data = request.json or {}
    sid = data.get('id', '')
    excel_path = data.get('excel_path', '')

    sheets = _load_kdocs_sheets()
    sheet_cfg = None
    for s in sheets:
        if s['id'] == sid:
            sheet_cfg = s
            break

    if not sheet_cfg:
        return jsonify({'error': '未找到该在线表格配置'}), 404

    api_token = sheet_cfg.get('api_token', '')
    webhook_url = sheet_cfg.get('webhook_url', '')
    batch_size = sheet_cfg.get('batch_size', 3)

    if not api_token or not webhook_url:
        return jsonify({'error': 'API_TOKEN 或 WEBHOOK_URL 未配置'}), 400

    # 如果传了excel_path则优先使用，否则用配置中的
    if not excel_path:
        excel_path = sheet_cfg.get('excel_path', '')
    if not excel_path:
        return jsonify({'error': '未指定本地Excel文件路径'}), 400

    if not os.path.exists(excel_path):
        return jsonify({'error': f'文件不存在: {excel_path}'}), 400

    # 解析webhook URL
    parsed = urlparse(webhook_url)
    host = parsed.hostname or "www.kdocs.cn"
    path = parsed.path

    # 测试连接
    try:
        conn = http.client.HTTPSConnection(host, context=ssl._create_unverified_context(), timeout=15)
        payload = json.dumps({"Context": {"argv": {"action": "info"}}}, ensure_ascii=False)
        headers = {"Content-Type": "application/json", "AirScript-Token": api_token}
        conn.request("POST", path, payload.encode("utf-8"), headers)
        res = conn.getresponse()
        result = json.loads(res.read().decode("utf-8"))
        conn.close()
        if result.get("error"):
            return jsonify({'error': f'Webhook连接异常: {result["error"]}'}), 500
    except Exception as e:
        return jsonify({'error': f'Webhook连接失败: {str(e)}'}), 500

    # 读取Excel
    try:
        df = pd.read_excel(excel_path)
        columns = list(df.columns)
        rows = []
        for _, row in df.iterrows():
            row_data = {}
            for col in columns:
                val = row[col]
                if pd.isna(val):
                    row_data[col] = ""
                elif isinstance(val, float) and val == int(val):
                    row_data[col] = int(val)
                else:
                    row_data[col] = str(val)
            rows.append(row_data)
    except Exception as e:
        return jsonify({'error': f'读取Excel失败: {str(e)}'}), 500

    # 分批写入
    total_batches = (len(rows) + batch_size - 1) // batch_size
    success_count = 0
    fail_count = 0
    details = []

    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        batch_num = i // batch_size + 1
        try:
            conn = http.client.HTTPSConnection(host, context=ssl._create_unverified_context(), timeout=30)
            pl = json.dumps({"Context": {"argv": {"columns": columns, "rows": batch}}}, ensure_ascii=False)
            headers = {"Content-Type": "application/json", "AirScript-Token": api_token}
            conn.request("POST", path, pl.encode("utf-8"), headers)
            resp = conn.getresponse()
            result = json.loads(resp.read().decode("utf-8"))
            conn.close()

            error = result.get("error", "")
            if error:
                fail_count += len(batch)
                details.append(f"批{batch_num}: 失败 - {error}")
                continue

            script_result = result.get("data", {}).get("result", "")
            try:
                ret = script_result
                if isinstance(ret, str) and ret and ret != "[Undefined]":
                    ret = json.loads(ret)
                if isinstance(ret, dict) and ret.get("success"):
                    success_count += ret.get("writeCount", len(batch))
                    details.append(f"批{batch_num}: 成功 - 写入{ret.get('writeCount','?')}行 (起始行{ret.get('startRow','?')})")
                else:
                    fail_count += len(batch)
                    details.append(f"批{batch_num}: 脚本返回错误 - {str(ret)[:100]}")
            except Exception as e:
                fail_count += len(batch)
                details.append(f"批{batch_num}: 解析异常 - {str(e)}")

        except Exception as e:
            fail_count += len(batch)
            details.append(f"批{batch_num}: 请求异常 - {str(e)}")

        if i + batch_size < len(rows):
            time.sleep(1.5)

    return jsonify({
        'message': f'推送完成：成功 {success_count} 行，失败 {fail_count} 行',
        'success_count': success_count,
        'fail_count': fail_count,
        'total_rows': len(rows),
        'details': details
    })


@app.route('/api/kdocs-push-batch', methods=['POST'])
def push_to_kdocs_batch():
    """一键推送：只推送与在线表格名称匹配的本地Excel文件"""
    data = request.json or {}
    folder_path = data.get('folder_path', '').strip()

    if not folder_path or not os.path.isdir(folder_path):
        return jsonify({'error': '文件夹路径无效'}), 400

    sheets = _load_kdocs_sheets()
    if not sheets:
        return jsonify({'error': '暂无在线表格配置'}), 400

    # 扫描文件夹下所有xlsx文件
    local_files = {}
    for f in os.listdir(folder_path):
        if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$'):
            local_files[f.lower()] = os.path.join(folder_path, f)

    results = []
    for s in sheets:
        # 模糊匹配：去掉扩展名后进行子串匹配（双向包含）
        online_name = s['name'].replace('.xlsx', '').replace('.xls', '').lower()
        matched_file = None
        matched_name = None
        for lf_name, lf_path in local_files.items():
            lf_base = lf_name.replace('.xlsx', '').replace('.xls', '')
            # 双向子串匹配：AAAB匹配AAABB
            if online_name and lf_base and (online_name in lf_base or lf_base in online_name):
                matched_file = lf_path
                matched_name = lf_name
                break

        if not matched_file:
            continue  # 未匹配的跳过，不推送也不报错

        if not s.get('api_token') or not s.get('webhook_url'):
            results.append({'id': s['id'], 'name': s['name'], 'file': matched_name, 'status': 'skip', 'message': 'API_TOKEN或WEBHOOK_URL未配置'})
            continue

        # 调用推送逻辑
        push_result = _do_push(s, matched_file)
        results.append({
            'id': s['id'],
            'name': s['name'],
            'file': matched_name,
            'status': 'ok' if push_result.get('fail_count', 0) == 0 else 'partial',
            'success_count': push_result.get('success_count', 0),
            'fail_count': push_result.get('fail_count', 0),
            'message': push_result.get('message', '')
        })

    return jsonify({'results': results, 'total_matched': len(results)})


def _do_push(sheet_cfg, excel_path):
    """执行单个推送（内部函数）"""
    api_token = sheet_cfg.get('api_token', '')
    webhook_url = sheet_cfg.get('webhook_url', '')
    batch_size = sheet_cfg.get('batch_size', 3)

    parsed = urlparse(webhook_url)
    host = parsed.hostname or "www.kdocs.cn"
    path = parsed.path

    # 读取Excel
    try:
        df = pd.read_excel(excel_path)
        columns = list(df.columns)
        rows = []
        for _, row in df.iterrows():
            row_data = {}
            for col in columns:
                val = row[col]
                if pd.isna(val):
                    row_data[col] = ""
                elif isinstance(val, float) and val == int(val):
                    row_data[col] = int(val)
                else:
                    row_data[col] = str(val)
            rows.append(row_data)
    except Exception as e:
        return {'success_count': 0, 'fail_count': 0, 'message': f'读取Excel失败: {str(e)}'}

    success_count = 0
    fail_count = 0

    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        try:
            conn = http.client.HTTPSConnection(host, context=ssl._create_unverified_context(), timeout=30)
            pl = json.dumps({"Context": {"argv": {"columns": columns, "rows": batch}}}, ensure_ascii=False)
            headers = {"Content-Type": "application/json", "AirScript-Token": api_token}
            conn.request("POST", path, pl.encode("utf-8"), headers)
            resp = conn.getresponse()
            result = json.loads(resp.read().decode("utf-8"))
            conn.close()

            error = result.get("error", "")
            if error:
                fail_count += len(batch)
                continue

            script_result = result.get("data", {}).get("result", "")
            ret = script_result
            if isinstance(ret, str) and ret and ret != "[Undefined]":
                try:
                    ret = json.loads(ret)
                except:
                    pass
            if isinstance(ret, dict) and ret.get("success"):
                success_count += ret.get("writeCount", len(batch))
            else:
                fail_count += len(batch)
        except:
            fail_count += len(batch)

        if i + batch_size < len(rows):
            time.sleep(1.5)

    msg = f'成功 {success_count} 行，失败 {fail_count} 行'
    return {'success_count': success_count, 'fail_count': fail_count, 'message': msg}


@app.route('/api/kdocs-airscript-code', methods=['GET'])
def get_airscript_code():
    """获取airscript_code.js脚本内容"""
    code_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'airscript_code.js')
    if not os.path.exists(code_path):
        return jsonify({'error': 'airscript_code.js 文件不存在'}), 404
    try:
        with open(code_path, 'r', encoding='utf-8') as f:
            code = f.read()
        return jsonify({'code': code})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/kdocs-airscript-code', methods=['PUT'])
def save_airscript_code():
    """保存airscript_code.js脚本内容"""
    data = request.json or {}
    code = data.get('code', '')
    code_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'airscript_code.js')
    try:
        with open(code_path, 'w', encoding='utf-8') as f:
            f.write(code)
        return jsonify({'message': '保存成功'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/kdocs-folder-scan', methods=['POST'])
def scan_folder():
    """扫描文件夹，列出所有Excel文件"""
    data = request.json or {}
    folder_path = data.get('folder_path', '').strip()
    if not folder_path or not os.path.isdir(folder_path):
        return jsonify({'error': '文件夹路径无效'}), 400

    files = []
    for f in os.listdir(folder_path):
        if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$'):
            fpath = os.path.join(folder_path, f)
            fsize = os.path.getsize(fpath)
            files.append({'name': f, 'path': fpath, 'size': fsize})

    return jsonify({'files': files, 'count': len(files)})


# ============ 数据标准化模板 API ============

def _nz_template_path(name):
    """安全获取数据标准化模板路径"""
    safe_name = re.sub(r'[^\w\u4e00-\u9fff\-\.]', '_', name)
    return os.path.join(NZ_TEMPLATES_DIR, f"{safe_name}.json")


def _nz_resolve_formula_str(formula_str, stats_data):
    """解析并计算单个公式字符串的值
    格式: {{fileIdx:L1:entryName:col/metric}} 或 {{fileIdx::entryName/metric}}
    entryName直接匹配entry.name（如"重点项"、"子项·重点项·核心项"等）
    兼容旧格式: {{fileIdx:L1:L2:L3/指标}} 或 {{fileIdx:总合计/指标}}
    """
    m = re.match(r'^\{\{(.+?)\}\}$', formula_str.strip())
    if not m:
        # 尝试匹配 {{...}}（可能在表达式内）
        m = re.search(r'\{\{(.+?)\}\}', formula_str.strip())
        if not m:
            return {'ok': False, 'value': '格式错误'}
    inner = m.group(1).strip()
    slash_idx = inner.rfind('/')
    if slash_idx < 0:
        return {'ok': False, 'value': '格式错误(无/指标)'}
    metric = inner[slash_idx + 1:].strip()
    path = inner[:slash_idx].strip()
    parts = path.split(':')
    parts = [p.strip() for p in parts]

    file_idx = None
    try:
        file_idx = int(parts[0])
    except (ValueError, IndexError):
        return {'ok': False, 'value': '文件序号无效'}

    if file_idx < 1:
        return {'ok': False, 'value': '文件序号无效'}

    fi_str = str(file_idx)
    if fi_str not in stats_data:
        return {'ok': False, 'value': '未匹配(文件不存在)'}

    fd = stats_data[fi_str]
    entries = fd.get('entries', [])
    total = fd.get('total', {})
    sum_col = fd.get('sumCol', '')

    # 解析各层级
    l1, col = '', ''
    entry_name = ''

    # 总合计
    if len(parts) == 2 and parts[1] == '总合计':
        return _nz_resolve_metric(total, '', metric, sum_col, None, None, fd)

    if len(parts) == 2:
        # {{1:entryName/指标}}
        entry_name = parts[1]
    elif len(parts) == 3:
        # {{1:L1:entryName/指标}} 或 {{1::entryName/指标}}
        l1 = parts[1]
        entry_name = parts[2]
    elif len(parts) == 4:
        # {{1:L1:entryName:col/指标}}
        l1 = parts[1]
        entry_name = parts[2]
        col = parts[3]
    elif len(parts) >= 5:
        # 旧多级格式兼容: {{fileIdx:L1:L2:L3:...:col/metric}}
        # 合并L2..L(N-1)为entry_name
        l1 = parts[1]
        remaining = parts[2:]
        # 最后一段如果包含.则为列名
        if remaining and '.' in remaining[-1]:
            col = remaining[-1]
            entry_name = ' · '.join(remaining[:-1])
        else:
            entry_name = ' · '.join(remaining)

    # 解析附加列.值
    ac_col, ac_val = None, None
    if col and '.' in col:
        dp = col.split('.', 1)
        ac_col, ac_val = dp[0], dp[1]

    # L1合计
    if l1 and entry_name == '合计':
        l1_total = None
        for e in entries:
            if e.get('isL1Total') and e.get('l1Name') == l1:
                l1_total = e
                break
        if not l1_total:
            return {'ok': False, 'value': '未匹配'}
        return _nz_resolve_metric(l1_total, col, metric, sum_col, ac_col, ac_val, fd)

    # 总合计（entry_name='总合计'）
    if entry_name == '总合计' or (not l1 and entry_name == '总合计'):
        return _nz_resolve_metric(total, col, metric, sum_col, ac_col, ac_val, fd)

    # 按 entry.name 直接匹配
    entry = None
    if l1:
        for e in entries:
            if e.get('isGroup') and e.get('l1Name') == l1 and e.get('name', '') == entry_name:
                entry = e
                break
        # 尝试L1合计
        if not entry:
            for e in entries:
                if e.get('isL1Total') and e.get('l1Name') == l1:
                    entry = e
                    break
    else:
        for e in entries:
            if e.get('isGroup') and not e.get('l1Name') and e.get('name', '') == entry_name:
                entry = e
                break

    if not entry:
        return {'ok': False, 'value': '未匹配'}

    return _nz_resolve_metric(entry, col, metric, sum_col, ac_col, ac_val, fd)


def _nz_resolve_metric(entry, col, metric, sum_col, ac_col, ac_val, fd):
    """解析指标值"""
    if metric == '数量':
        if ac_col and ac_val:
            ac_data = entry.get('acData', {})
            tc = ac_data.get(ac_col, {})
            return {'ok': True, 'value': tc.get(ac_val, 0)}
        return {'ok': True, 'value': entry.get('count', 0)}
    if metric == '占比':
        pct = entry.get('pct', '0')
        try:
            return {'ok': True, 'value': float(pct)}
        except:
            return {'ok': False, 'value': '占比解析失败'}
    if metric == '求和':
        target_col = col or sum_col
        if not target_col:
            return {'ok': False, 'value': '未匹配(无求和列)'}
        s = entry.get('sum')
        if s is not None:
            try:
                return {'ok': True, 'value': float(s)}
            except:
                pass
        return {'ok': False, 'value': '未匹配(求和)'}
    if metric == '均值':
        target_col = col or sum_col
        if not target_col:
            return {'ok': False, 'value': '未匹配(无求和列)'}
        s = entry.get('sum')
        c = entry.get('count', 0)
        if s is not None and c > 0:
            try:
                return {'ok': True, 'value': float(s) / c}
            except:
                pass
        return {'ok': True, 'value': 0}
    return {'ok': False, 'value': '未知指标'}


@app.route('/api/nz-templates', methods=['GET'])
def list_nz_templates():
    """列出所有数据标准化模板"""
    templates = []
    if os.path.exists(NZ_TEMPLATES_DIR):
        for fname in os.listdir(NZ_TEMPLATES_DIR):
            if not fname.endswith('.json'):
                continue
            fpath = os.path.join(NZ_TEMPLATES_DIR, fname)
            try:
                with open(fpath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                templates.append({
                    'name': data.get('name', fname[:-5]),
                    'savedAt': data.get('savedAt', 0),
                    'sheetCount': data.get('sheetCount', 0)
                })
            except:
                pass
    templates.sort(key=lambda t: t.get('savedAt', 0), reverse=True)
    return jsonify({'templates': templates})


@app.route('/api/nz-templates', methods=['POST'])
def save_nz_template():
    """保存数据标准化模板"""
    data = request.json or {}
    name = (data.get('name') or '').strip()
    file_data = data.get('fileData', '')
    if not name:
        return jsonify({'error': '模板名称不能为空'}), 400
    if not file_data:
        return jsonify({'error': '模板数据不能为空'}), 400

    # 验证base64是有效的Excel
    try:
        raw = base64.b64decode(file_data)
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp.write(raw)
        tmp.close()
        wb = load_workbook(tmp.name)
        sheet_count = len(wb.sheetnames)
        wb.close()
        os.unlink(tmp.name)
    except Exception as e:
        return jsonify({'error': f'模板文件无效: {str(e)}'}), 400

    template_data = {
        'name': name,
        'fileData': file_data,
        'savedAt': datetime.now().timestamp() * 1000,
        'sheetCount': sheet_count
    }
    fpath = _nz_template_path(name)
    with open(fpath, 'w', encoding='utf-8') as f:
        json.dump(template_data, f, ensure_ascii=False)
    return jsonify({'message': '模板已保存', 'name': name})


@app.route('/api/nz-templates/<path:name>', methods=['GET'])
def get_nz_template(name):
    """获取指定数据标准化模板"""
    fpath = _nz_template_path(name)
    if not os.path.exists(fpath):
        return jsonify({'error': '模板不存在'}), 404
    with open(fpath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return app.response_class(
        response=json.dumps(data, ensure_ascii=False),
        status=200,
        mimetype='application/json'
    )


@app.route('/api/nz-templates/<path:name>', methods=['DELETE'])
def delete_nz_template(name):
    """删除指定数据标准化模板"""
    fpath = _nz_template_path(name)
    if not os.path.exists(fpath):
        return jsonify({'error': '模板不存在'}), 404
    os.remove(fpath)
    return jsonify({'message': '模板已删除'})


@app.route('/api/nz-fill', methods=['POST'])
def nz_fill_template():
    """填充数据标准化模板：解析公式并替换值，返回填充后的Excel文件
    使用原始模板文件保留格式，仅替换公式单元格的值。
    """
    data = request.json or {}
    template_b64 = data.get('templateData', '')
    stats_data = data.get('statsData', {})
    cell_edits = data.get('cellEdits', [])

    if not template_b64:
        return jsonify({'error': '缺少模板数据'}), 400

    # 解码base64模板
    try:
        raw = base64.b64decode(template_b64)
    except:
        return jsonify({'error': '模板数据解码失败'}), 400

    # 写入临时文件并用openpyxl打开（保留原格式）
    tmp_in = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_in.write(raw)
    tmp_in.close()
    tmp_out = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_out.close()

    try:
        wb = load_workbook(tmp_in.name)
    except Exception as e:
        os.unlink(tmp_in.name)
        os.unlink(tmp_out.name)
        return jsonify({'error': f'读取模板失败: {str(e)}'}), 500

    # 1. 先应用前端单元格编辑（保留格式，只改值）
    for edit in cell_edits:
        si = edit.get('sheet', 0)
        r = edit.get('row', 0)
        c = edit.get('col', 0)
        val = edit.get('value', '')
        if si < 0 or si >= len(wb.sheetnames):
            continue
        ws = wb.worksheets[si]
        cell = ws.cell(row=r + 1, column=c + 1)  # openpyxl是1-based
        # 保留原有格式，只替换值
        try:
            # 尝试转为数字
            num_val = float(val) if val and '.' in val else (int(val) if val and val.lstrip('-').isdigit() else None)
        except (ValueError, TypeError, AttributeError):
            num_val = None
        if num_val is not None:
            cell.value = num_val
        else:
            cell.value = val

    # 2. 应用单元格格式
    cell_formats = data.get('cellFormats', [])
    for fmt_item in cell_formats:
        si = fmt_item.get('sheet', 0)
        r = fmt_item.get('row', 0)
        c = fmt_item.get('col', 0)
        fmt = fmt_item.get('fmt', {})
        if si < 0 or si >= len(wb.sheetnames):
            continue
        ws = wb.worksheets[si]
        cell = ws.cell(row=r + 1, column=c + 1)
        if fmt.get('bold') or fmt.get('italic') or fmt.get('fontSize') or fmt.get('fontName'):
            cell.font = Font(
                bold=fmt.get('bold', False),
                italic=fmt.get('italic', False),
                size=fmt.get('fontSize'),
                name=fmt.get('fontName')
            )
        if fmt.get('align'):
            align_map = {'left': 'left', 'center': 'center', 'right': 'right'}
            cell.alignment = Alignment(horizontal=align_map.get(fmt['align'], 'left'))

    # 3. 扫描公式并替换值（支持多公式表达式、单元格引用和混合文本）
    formula_pattern = re.compile(r'\{\{(.+?)\}\}')
    # 单元格引用正则：A1, B2, AA10 等
    cell_ref_pattern = re.compile(r'\b([A-Z]{1,3})(\d{1,5})\b')
    fill_count = 0
    fail_count = 0

    # 先构建每个sheet的值缓存（用于单元格引用解析）
    def get_cell_value(ws_obj, col_str, row_str):
        """根据列字母和行号获取单元格值"""
        try:
            col_idx = 0
            for ch in col_str:
                col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
            row_idx = int(row_str)
            if row_idx < 1 or col_idx < 1:
                return None
            c = ws_obj.cell(row=row_idx, column=col_idx)
            if c.value is None:
                return 0
            return c.value
        except:
            return None

    def resolve_cell_refs_in_expr(expr_str, ws_obj, visited=None):
        """解析表达式中的单元格引用，替换为实际值"""
        if visited is None:
            visited = set()
        def replace_ref(m):
            col_s, row_s = m.group(1), m.group(2)
            ref_key = f"{col_s}{row_s}"
            if ref_key in visited:
                return 'NaN'
            visited.add(ref_key)
            val = get_cell_value(ws_obj, col_s, row_s)
            if val is None:
                return 'NaN'
            val_str = str(val).strip()
            # 递归解析：如果引用的单元格也是表达式
            if val_str.startswith('='):
                inner = val_str[1:]
                # 先替换 {{}}
                inner = formula_pattern.sub(lambda fm: str(_nz_resolve_formula_str(fm.group(0), stats_data).get('value', 'NaN')), inner)
                # 再替换引用
                inner = resolve_cell_refs_in_expr(inner, ws_obj, visited)
                try:
                    safe = re.sub(r'[^0-9+\-*/().eE\s]', '', inner)
                    result = eval(safe)
                    if isinstance(result, (int, float)):
                        return str(result)
                    return 'NaN'
                except:
                    return 'NaN'
            try:
                num = float(val)
                if '.' not in str(val) and isinstance(val, int):
                    return str(val)
                return str(num)
            except (ValueError, TypeError):
                return 'NaN'
        return cell_ref_pattern.sub(replace_ref, expr_str)

    # 解析范围引用中的所有单元格值
    def resolve_range_values(range_str, ws_obj, visited=None):
        """解析 A1:B5 或 A1,B2,C3 格式，返回数值列表"""
        if visited is None:
            visited = set()
        values = []
        parts = [p.strip() for p in range_str.split(',')]
        range_re = re.compile(r'^([A-Z]{1,3})(\d{1,5}):([A-Z]{1,3})(\d{1,5})$')
        cell_re = re.compile(r'^([A-Z]{1,3})(\d{1,5})$')
        for part in parts:
            rm = range_re.match(part)
            if rm:
                c1 = 0
                for ch in rm.group(1):
                    c1 = c1 * 26 + (ord(ch) - ord('A') + 1)
                r1 = int(rm.group(2))
                c2 = 0
                for ch in rm.group(3):
                    c2 = c2 * 26 + (ord(ch) - ord('A') + 1)
                r2 = int(rm.group(4))
                min_r, max_r = min(r1, r2), max(r1, r2)
                min_c, max_c = min(c1, c2), max(c1, c2)
                for rr in range(min_r, max_r + 1):
                    for cc in range(min_c, max_c + 1):
                        ref_key = f"{rr},{cc}"
                        if ref_key in visited:
                            continue
                        visited.add(ref_key)
                        cv = ws_obj.cell(row=rr, column=cc).value
                        if cv is not None:
                            try:
                                v = float(cv)
                                values.append(v)
                            except (ValueError, TypeError):
                                pass
            else:
                cm = cell_re.match(part)
                if cm:
                    col_s, row_s = cm.group(1), cm.group(2)
                    ref_key = f"{col_s}{row_s}"
                    if ref_key not in visited:
                        visited.add(ref_key)
                        val = get_cell_value(ws_obj, col_s, row_s)
                        if val is not None:
                            try:
                                values.append(float(val))
                            except (ValueError, TypeError):
                                pass
        return values

    # 替换 SUM(...)/AVG(...) 函数
    func_pattern = re.compile(r'\b(SUM|AVG)\s*\(([^)]+)\)', re.IGNORECASE)

    def resolve_funcs(expr_str, ws_obj, visited=None):
        """解析 SUM/AVG 函数，替换为计算结果"""
        if visited is None:
            visited = set()
        def replace_func(m):
            fn = m.group(1).upper()
            args = m.group(2)
            vals = resolve_range_values(args, ws_obj, visited)
            if not vals:
                return 'NaN'
            if fn == 'SUM':
                return str(sum(vals))
            elif fn == 'AVG':
                return str(sum(vals) / len(vals))
            return 'NaN'
        return func_pattern.sub(replace_func, expr_str)

    for ws in wb.worksheets:
        # 遍历所有单元格
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_str = str(cell.value).strip()

                # 判断是否包含公式
                has_formula = formula_pattern.search(cell_str) or cell_str.startswith('=')
                if not has_formula:
                    continue

                # 判断公式模式
                if cell_str.startswith('='):
                    # 表达式模式：=A1+B2 或 ={{...}}+A1 或混合
                    expr = cell_str[1:]  # 去掉 = 前缀
                    all_ok = True

                    # 1. 替换 {{...}} 公式
                    def resolve_expr(match):
                        nonlocal all_ok
                        result = _nz_resolve_formula_str(match.group(0), stats_data)
                        if result['ok']:
                            return str(result['value'])
                        all_ok = False
                        return 'NaN'

                    expr = formula_pattern.sub(resolve_expr, expr)
                    if not all_ok:
                        cell.value = '公式解析失败'
                        fail_count += 1
                        continue

                    # 2. 替换 SUM(...)/AVG(...) 函数
                    visited = set()
                    self_col = get_column_letter(cell.column)
                    self_row = str(cell.row)
                    visited.add(f"{self_col}{self_row}")
                    expr = resolve_funcs(expr, ws, visited)
                    if 'NaN' in expr:
                        cell.value = '函数解析失败'
                        fail_count += 1
                        continue

                    # 3. 替换单元格引用
                    expr = resolve_cell_refs_in_expr(expr, ws, visited)
                    if 'NaN' in expr:
                        cell.value = '引用解析失败'
                        fail_count += 1
                        continue

                    # 4. 安全计算
                    safe_expr = re.sub(r'[^0-9+\-*/().eE\s]', '', expr)
                    try:
                        num_result = eval(safe_expr)
                        if isinstance(num_result, (int, float)) and str(num_result) != 'nan':
                            cell.value = num_result
                            fill_count += 1
                        else:
                            cell.value = '计算错误'
                            fail_count += 1
                    except:
                        cell.value = '表达式错误'
                        fail_count += 1

                elif re.match(r'^\{\{(.+?)\}\}$', cell_str):
                    # 单公式模式（向后兼容）
                    result = _nz_resolve_formula_str(cell_str, stats_data)
                    if result['ok']:
                        val = result['value']
                        cell.value = val
                        fill_count += 1
                    else:
                        cell.value = str(result['value'])
                        fail_count += 1

                else:
                    # 混合文本模式：替换所有 {{...}} 为数值
                    any_fail = False

                    def resolve_text(match):
                        nonlocal any_fail
                        result = _nz_resolve_formula_str(match.group(0), stats_data)
                        if result['ok']:
                            return str(result['value'])
                        any_fail = True
                        return match.group(0)

                    final_val = formula_pattern.sub(resolve_text, cell_str)
                    cell.value = final_val
                    if any_fail:
                        fail_count += 1
                    else:
                        fill_count += 1

    # 4. 应用数值格式（小数位数 + 百分比）
    # 构建 (sheet,row,col) -> fmt 的快速查找表
    fmt_lookup = {}
    for fmt_item in cell_formats:
        key = (fmt_item.get('sheet', 0), fmt_item.get('row', 0), fmt_item.get('col', 0))
        fmt_lookup[key] = fmt_item.get('fmt', {})

    for ws_idx, ws in enumerate(wb.worksheets):
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                key = (ws_idx, cell.row - 1, cell.column - 1)
                fmt = fmt_lookup.get(key, {})
                decimal = fmt.get('decimal', None)
                isPercent = fmt.get('percent', False)
                if decimal is None and not isPercent:
                    continue
                try:
                    v = float(cell.value)
                except (ValueError, TypeError):
                    continue
                if isPercent:
                    v = v * 100
                d = decimal if decimal is not None else 2
                factor = 10 ** d
                rounded = round(v * factor) / factor
                if isPercent:
                    cell.value = f"{rounded:.{d}f}%"
                else:
                    cell.value = round(v, d) if d > 0 else int(rounded)
                # 百分比用百分比数字格式
                if isPercent:
                    cell.number_format = '0%' if d == 0 else f'0.{"0" * d}%'
                elif d > 0:
                    cell.number_format = f'0.{"0" * d}'

    try:
        wb.save(tmp_out.name)
        wb.close()
    except Exception as e:
        wb.close()
        os.unlink(tmp_in.name)
        os.unlink(tmp_out.name)
        return jsonify({'error': f'保存失败: {str(e)}'}), 500

    # 发送文件
    result = send_file(
        tmp_out.name,
        as_attachment=True,
        download_name='填充结果.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # 清理（使用after_request延迟清理）
    @result.call_on_close
    def cleanup():
        try:
            os.unlink(tmp_in.name)
            os.unlink(tmp_out.name)
        except:
            pass

    return result


# ================================================================
#  邮件发送模块（从 email-tool 整合）
# ================================================================

# ---- 邮件数据目录 ----
EMAIL_DATA_DIR = os.path.join(DATA_DIR, 'email')
os.makedirs(EMAIL_DATA_DIR, exist_ok=True)
EMAIL_CONTACTS_FILE = os.path.join(EMAIL_DATA_DIR, 'contacts.json')
EMAIL_TEMPLATES_FILE = os.path.join(EMAIL_DATA_DIR, 'templates.json')
EMAIL_COOKIES_FILE = os.path.join(EMAIL_DATA_DIR, 'cookies.json')
EMAIL_UPLOAD_DIR = os.path.join(EMAIL_DATA_DIR, 'uploads')
os.makedirs(EMAIL_UPLOAD_DIR, exist_ok=True)
EMAIL_LOGIN_CREDS_FILE = os.path.join(EMAIL_DATA_DIR, 'login_creds.json')
TXT_VARS_DIR = os.path.join(EMAIL_DATA_DIR, 'txt_vars')
os.makedirs(TXT_VARS_DIR, exist_ok=True)

MAIL_CONFIG = {
    'username': 'wangy592@chinatelecom.cn',
    'password': 'wY0426!..',
    'auth_code': 'nblaelviyhpdegbh',
    'account': 'wangy592',
    'phone': '18081927229',
}

login_state = {
    'status': 'idle',
    'message': '',
    'code': None,
    'browser_open': False,
}


def _email_load_json(filepath, default=None):
    if default is None:
        default = []
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return default
    return default


def _email_save_json(filepath, data):
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _load_login_creds():
    data = _email_load_json(EMAIL_LOGIN_CREDS_FILE, default=None)
    if data and data.get('account'):
        return data
    return {'account': MAIL_CONFIG['account'], 'password': MAIL_CONFIG['password'], 'phone': MAIL_CONFIG['phone']}


def _save_login_creds(account, password, phone):
    _email_save_json(EMAIL_LOGIN_CREDS_FILE, {'account': account, 'password': password, 'phone': phone})


# ---- 收件人管理 ----
@app.route('/api/email/contacts', methods=['GET'])
def email_get_contacts():
    return jsonify({'success': True, 'data': _email_load_json(EMAIL_CONTACTS_FILE)})


@app.route('/api/email/contacts', methods=['POST'])
def email_add_contact():
    data = request.json
    contacts = _email_load_json(EMAIL_CONTACTS_FILE)
    contact = {
        'id': int(time.time() * 1000),
        'name': data.get('name', ''),
        'email': data.get('email', ''),
        'group': data.get('group', '默认分组')
    }
    contacts.append(contact)
    _email_save_json(EMAIL_CONTACTS_FILE, contacts)
    return jsonify({'success': True, 'data': contact})


@app.route('/api/email/contacts/<int:cid>', methods=['DELETE'])
def email_delete_contact(cid):
    contacts = [c for c in _email_load_json(EMAIL_CONTACTS_FILE) if c.get('id') != cid]
    _email_save_json(EMAIL_CONTACTS_FILE, contacts)
    return jsonify({'success': True})


@app.route('/api/email/contacts/<int:cid>', methods=['PUT'])
def email_update_contact(cid):
    data = request.json
    contacts = _email_load_json(EMAIL_CONTACTS_FILE)
    for c in contacts:
        if c.get('id') == cid:
            c['name'] = data.get('name', c['name'])
            c['email'] = data.get('email', c['email'])
            c['group'] = data.get('group', c.get('group', '默认分组'))
            break
    _email_save_json(EMAIL_CONTACTS_FILE, contacts)
    return jsonify({'success': True})


# ---- 分组管理 ----
@app.route('/api/email/groups', methods=['GET'])
def email_get_groups():
    contacts = _email_load_json(EMAIL_CONTACTS_FILE)
    groups = {}
    for c in contacts:
        g = c.get('group', '默认分组')
        if g not in groups:
            groups[g] = {'name': g, 'count': 0, 'contacts': []}
        groups[g]['count'] += 1
        groups[g]['contacts'].append({'id': c['id'], 'name': c['name'], 'email': c['email'], 'group': g})
    return jsonify({'success': True, 'data': list(groups.values())})


# ---- Cookie 验证 ----
@app.route('/api/email/login/check', methods=['GET'])
def email_check_login():
    cookies_data = _email_load_json(EMAIL_COOKIES_FILE, default=None)
    if not cookies_data:
        return jsonify({'success': True, 'logged_in': False})
    csrftoken_file = os.path.join(EMAIL_DATA_DIR, 'csrftoken.txt')
    if not os.path.exists(csrftoken_file):
        return jsonify({'success': True, 'logged_in': False})
    with open(csrftoken_file, 'r', encoding='utf-8') as f:
        csrftoken = f.read().strip()
    csrftoken = re.sub(r'[\u200b\u200c\u200d\ufeff\u00a0\s]', '', csrftoken)
    if not csrftoken:
        return jsonify({'success': True, 'logged_in': False})
    try:
        import requests as req
        session = req.Session()
        for c in cookies_data:
            session.cookies.set(c['name'], c['value'], domain=c.get('domain', ''))
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'csrftoken': csrftoken,
            'Referer': 'https://mail.chinatelecom.cn/mail/index.html',
        })
        resp = session.post('https://mail.chinatelecom.cn/w2/replay/getRandomNum', timeout=8)
        if resp.status_code == 200 and resp.json().get('code') == 0:
            return jsonify({'success': True, 'logged_in': True})
    except Exception:
        pass
    return jsonify({'success': True, 'logged_in': False})


# ---- txt 变量文件 ----
@app.route('/api/email/txt-vars', methods=['POST'])
def email_upload_txt_var():
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'success': False, 'message': '请选择文件'}), 400
    if not f.filename.endswith('.txt'):
        return jsonify({'success': False, 'message': '仅支持 .txt 文件'}), 400
    basename = os.path.splitext(f.filename)[0]
    content = f.read().decode('utf-8', errors='ignore')
    lines = [line.strip() for line in content.split('\n') if line.strip()]
    save_path = os.path.join(TXT_VARS_DIR, f'{basename}.json')
    _email_save_json(save_path, {'name': basename, 'values': lines})
    return jsonify({'success': True, 'data': {'name': basename, 'count': len(lines)}})


@app.route('/api/email/txt-vars', methods=['GET'])
def email_list_txt_vars():
    result = []
    for fname in os.listdir(TXT_VARS_DIR):
        if fname.endswith('.json'):
            data = _email_load_json(os.path.join(TXT_VARS_DIR, fname), default={})
            if data:
                result.append({'name': data.get('name', ''), 'count': len(data.get('values', []))})
    return jsonify({'success': True, 'data': result})


@app.route('/api/email/txt-vars/<name>', methods=['GET'])
def email_get_txt_var(name):
    fpath = os.path.join(TXT_VARS_DIR, f'{name}.json')
    data = _email_load_json(fpath, default=None)
    if not data:
        return jsonify({'success': False, 'message': '变量文件不存在'}), 404
    return jsonify({'success': True, 'data': data})


@app.route('/api/email/txt-vars/<name>', methods=['DELETE'])
def email_delete_txt_var(name):
    fpath = os.path.join(TXT_VARS_DIR, f'{name}.json')
    if os.path.exists(fpath):
        os.remove(fpath)
    return jsonify({'success': True})


# ---- 图片上传 ----
@app.route('/api/email/upload-image', methods=['POST'])
def email_upload_image():
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'success': False, 'message': '请选择图片'}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'):
        return jsonify({'success': False, 'message': '仅支持图片文件'}), 400
    fname = f'{int(time.time()*1000)}{ext}'
    fpath = os.path.join(EMAIL_UPLOAD_DIR, fname)
    f.save(fpath)
    return jsonify({'success': True, 'data': {'url': f'/email-uploads/{fname}', 'filename': fname}})


@app.route('/email-uploads/<path:filename>')
def email_serve_upload(filename):
    return send_from_directory(EMAIL_UPLOAD_DIR, filename)


# ---- 模板管理 ----
@app.route('/api/email/templates', methods=['GET'])
def email_get_templates():
    return jsonify({'success': True, 'data': _email_load_json(EMAIL_TEMPLATES_FILE)})


@app.route('/api/email/templates', methods=['POST'])
def email_add_template():
    data = request.json
    templates = _email_load_json(EMAIL_TEMPLATES_FILE)
    template = {
        'id': int(time.time() * 1000),
        'name': data.get('name', '未命名模板'),
        'subject': data.get('subject', ''),
        'body': data.get('body', ''),
        'to': data.get('to', []),
        'cc': data.get('cc', []),
        'batchMode': data.get('batchMode', False),
    }
    templates.append(template)
    _email_save_json(EMAIL_TEMPLATES_FILE, templates)
    return jsonify({'success': True, 'data': template})


@app.route('/api/email/templates/<int:tid>', methods=['DELETE'])
def email_delete_template(tid):
    templates = [t for t in _email_load_json(EMAIL_TEMPLATES_FILE) if t.get('id') != tid]
    _email_save_json(EMAIL_TEMPLATES_FILE, templates)
    return jsonify({'success': True})


@app.route('/api/email/templates/<int:tid>', methods=['PUT'])
def email_update_template(tid):
    data = request.json
    templates = _email_load_json(EMAIL_TEMPLATES_FILE)
    for t in templates:
        if t.get('id') == tid:
            t['name'] = data.get('name', t['name'])
            t['subject'] = data.get('subject', t['subject'])
            t['body'] = data.get('body', t['body'])
            if 'to' in data:
                t['to'] = data['to']
            if 'cc' in data:
                t['cc'] = data['cc']
            if 'batchMode' in data:
                t['batchMode'] = data['batchMode']
            break
    _email_save_json(EMAIL_TEMPLATES_FILE, templates)
    return jsonify({'success': True})


# ---- 邮件发送 ----
def _email_upload_attachment(session, csrftoken, file_path, filename):
    url = 'https://mail.chinatelecom.cn/w2/common/uploadFile'
    upload_headers = {
        'csrftoken': csrftoken,
        'Origin': 'https://mail.chinatelecom.cn',
        'Referer': 'https://mail.chinatelecom.cn/mail/index.html',
    }
    with open(file_path, 'rb') as f:
        file_data = f.read()
    # 手动构造 multipart，确保 Content-Disposition 中 filename 使用 UTF-8 编码
    # 避免 requests 库对中文文件名产生 =?UTF-8B?...?= 编码问题
    boundary = '----WebHostFormBoundary' + ''.join(random.choices('0123456789abcdef', k=16))
    body_parts = []
    body_parts.append(f'--{boundary}'.encode('utf-8'))
    # 对文件名进行 RFC 2047 编码
    try:
        filename.encode('ascii')
        disp_filename = filename
    except UnicodeEncodeError:
        b64_name = base64.b64encode(filename.encode('utf-8')).decode('ascii')
        disp_filename = f'=?UTF-8?B?{b64_name}?='
    cd = f'Content-Disposition: form-data; name="file"; filename="{disp_filename}"'
    body_parts.append(cd.encode('utf-8'))
    body_parts.append(b'Content-Type: application/octet-stream')
    body_parts.append(b'')
    body_parts.append(file_data)
    body_parts.append(f'--{boundary}--'.encode('utf-8'))
    body_bytes = b'\r\n'.join(body_parts)
    upload_headers['Content-Type'] = f'multipart/form-data; boundary={boundary}'
    resp = session.post(url, data=body_bytes, headers=upload_headers, timeout=60)
    if resp.status_code == 200:
        try:
            data = resp.json()
            if data.get('code') == 0 and data.get('data'):
                return data['data'][0].get('fileKey', '')
        except Exception:
            pass
    return None


@app.route('/api/email/send', methods=['POST'])
def email_send():
    import json as _json
    to_emails = _json.loads(request.form.get('to', '[]'))
    cc_emails = _json.loads(request.form.get('cc', '[]'))
    subject = request.form.get('subject', '')
    body = request.form.get('body', '')
    if not to_emails:
        return jsonify({'success': False, 'message': '请填写收件人'}), 400
    if not subject:
        return jsonify({'success': False, 'message': '请填写邮件主题'}), 400
    cookies_data = _email_load_json(EMAIL_COOKIES_FILE, default=None)
    if not cookies_data:
        return jsonify({'success': False, 'message': '请先登录邮箱（点击登录管理进行黑箱登录）'}), 400
    try:
        import requests as req
        session = req.Session()
        for c in cookies_data:
            session.cookies.set(c['name'], c['value'], domain=c.get('domain', ''))
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Referer': 'https://mail.chinatelecom.cn/mail/index.html',
        })
        csrftoken_file = os.path.join(EMAIL_DATA_DIR, 'csrftoken.txt')
        if not os.path.exists(csrftoken_file):
            return jsonify({'success': False, 'message': '登录会话已过期，请重新登录邮箱'}), 400
        with open(csrftoken_file, 'r', encoding='utf-8') as f:
            csrftoken = f.read().strip()
        csrftoken = re.sub(r'[\u200b\u200c\u200d\ufeff\u00a0\s]', '', csrftoken)
        if not csrftoken:
            return jsonify({'success': False, 'message': '登录会话已过期，请重新登录邮箱'}), 400
        session.headers.update({'csrftoken': csrftoken})
        random_resp = session.post('https://mail.chinatelecom.cn/w2/replay/getRandomNum', timeout=10)
        if random_resp.status_code != 200:
            return jsonify({'success': False, 'message': f'获取安全码失败 (HTTP {random_resp.status_code})'}), 500
        random_data = random_resp.json()
        if random_data.get('code') != 0:
            return jsonify({'success': False, 'message': f'获取安全码失败: {random_data.get("desc", "未知错误")}'}), 500
        security_code = random_data.get('data', '')
        html_body = body.replace('\n', '<br>') if '<' not in body else body
        if '<p>' not in html_body and '<br>' not in html_body and '<div' not in html_body:
            html_body = '<p>' + html_body + '</p>'
        content_b64 = base64.b64encode(html_body.encode('utf-8')).decode('utf-8')
        attachment_list, attachment_name_list, upload_errors = [], [], []
        uploaded_files = request.files.getlist('files')
        for f in uploaded_files:
            if f.filename:
                tmp_path = os.path.join(EMAIL_UPLOAD_DIR, f.filename)
                f.save(tmp_path)
                try:
                    file_key = _email_upload_attachment(session, csrftoken, tmp_path, f.filename)
                    if file_key:
                        attachment_list.append(file_key)
                        # 附件名使用 RFC 2047 编码，确保中文正确显示
                        try:
                            f.filename.encode('ascii')
                            safe_name = f.filename
                        except UnicodeEncodeError:
                            b64 = base64.b64encode(f.filename.encode('utf-8')).decode('ascii')
                            safe_name = f'=?UTF-8?B?{b64}?='
                        attachment_name_list.append(safe_name)
                    else:
                        upload_errors.append(f"附件 '{f.filename}' 上传失败")
                except Exception as e:
                    upload_errors.append(f"附件 '{f.filename}' 上传异常: {str(e)}")
                finally:
                    try: os.remove(tmp_path)
                    except OSError: pass
        attachment_list_str = ','.join(attachment_list)
        attachment_name_list_str = ','.join(attachment_name_list)
        send_data = {
            'from': MAIL_CONFIG['username'], 'to': ','.join(to_emails), 'cc': ','.join(cc_emails),
            'bcc': '', 'fast': '0', 'content': content_b64, 'contentType': '1',
            'subject': subject, 'attachmentList': attachment_list_str,
            'attachmentNameList': attachment_name_list_str, 'dnt': '0',
            'action': 'send', 'sendMode': '0', 'saveSended': '1',
            'securityDestroy': '0', 'acceptSmsphones': '', 'acceptSmsKey': '',
            'securityCode': security_code,
        }
        encoded_body = urllib.parse.urlencode(send_data, encoding='utf-8')
        send_resp = session.post('https://mail.chinatelecom.cn/w2/mail/sendMail',
            data=encoded_body.encode('utf-8'),
            headers={'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'}, timeout=30)
        if send_resp.status_code != 200:
            return jsonify({'success': False, 'message': f'发信接口返回错误 (HTTP {send_resp.status_code})'}), 500
        result = send_resp.json()
        if result.get('code') == 0:
            msg = f'邮件已成功发送至 {len(to_emails)} 位收件人'
            if attachment_list and uploaded_files and len(attachment_list) < len(uploaded_files):
                msg += '（部分附件上传失败）'
            if upload_errors:
                msg += f'。附件问题: {"; ".join(upload_errors)}'
            return jsonify({'success': True, 'message': msg})
        else:
            return jsonify({'success': False, 'message': f'发送失败: {result.get("desc", "未知错误")}'}), 500
    except req.exceptions.ConnectionError:
        return jsonify({'success': False, 'message': '网络连接失败'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': f'发送失败: {str(e)}'}), 500


# ---- 批量个性化发信 ----
@app.route('/api/email/batch-send', methods=['POST'])
def email_batch_send():
    import json as _json
    items_json = request.form.get('items', '[]')
    cc_emails = _json.loads(request.form.get('cc', '[]'))
    try:
        items = _json.loads(items_json)
    except Exception:
        return jsonify({'success': False, 'message': '发信数据格式错误'}), 400
    if not items:
        return jsonify({'success': False, 'message': '请添加至少一位收件人'}), 400
    cookies_data = _email_load_json(EMAIL_COOKIES_FILE, default=None)
    if not cookies_data:
        return jsonify({'success': False, 'message': '请先登录邮箱'}), 400
    try:
        import requests as req
        session = req.Session()
        for c in cookies_data:
            session.cookies.set(c['name'], c['value'], domain=c.get('domain', ''))
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Referer': 'https://mail.chinatelecom.cn/mail/index.html',
        })
        csrftoken_file = os.path.join(EMAIL_DATA_DIR, 'csrftoken.txt')
        if not os.path.exists(csrftoken_file):
            return jsonify({'success': False, 'message': '登录会话已过期，请重新登录邮箱'}), 400
        with open(csrftoken_file, 'r', encoding='utf-8') as f:
            csrftoken = f.read().strip()
        csrftoken = re.sub(r'[\u200b\u200c\u200d\ufeff\u00a0\s]', '', csrftoken)
        if not csrftoken:
            return jsonify({'success': False, 'message': '登录会话已过期，请重新登录邮箱'}), 400
        session.headers.update({'csrftoken': csrftoken})
        common_tmp_files = []
        uploaded_files = request.files.getlist('files')
        for f in uploaded_files:
            if f.filename:
                tmp_path = os.path.join(EMAIL_UPLOAD_DIR, f'common_{int(time.time()*1000)}_{f.filename}')
                f.save(tmp_path)
                common_tmp_files.append((tmp_path, f.filename))
        success_count, fail_list = 0, []
        for idx, item in enumerate(items):
            to_email = item.get('to', '')
            subject = item.get('subject', '')
            body = item.get('body', '')
            if not to_email or not subject:
                fail_list.append(f"{to_email}: 缺少收件人或主题")
                continue
            random_resp = session.post('https://mail.chinatelecom.cn/w2/replay/getRandomNum', timeout=10)
            if random_resp.status_code != 200:
                fail_list.append(f"{to_email}: 获取安全码失败")
                continue
            random_data = random_resp.json()
            if random_data.get('code') != 0:
                fail_list.append(f"{to_email}: 获取安全码失败")
                continue
            security_code = random_data.get('data', '')
            html_body = body.replace('\n', '<br>') if '<' not in body else body
            if '<p>' not in html_body and '<br>' not in html_body and '<div' not in html_body:
                html_body = '<p>' + html_body + '</p>'
            content_b64 = base64.b64encode(html_body.encode('utf-8')).decode('utf-8')
            common_keys, common_names = [], []
            for tmp_path, orig_name in common_tmp_files:
                try:
                    file_key = _email_upload_attachment(session, csrftoken, tmp_path, orig_name)
                    if file_key:
                        common_keys.append(file_key)
                        try:
                            orig_name.encode('ascii')
                            safe_name = orig_name
                        except UnicodeEncodeError:
                            b64 = base64.b64encode(orig_name.encode('utf-8')).decode('ascii')
                            safe_name = f'=?UTF-8?B?{b64}?='
                        common_names.append(safe_name)
                except Exception:
                    pass
            per_files = request.files.getlist(f'files_{idx}')
            per_keys, per_names = [], []
            for f in per_files:
                if f.filename:
                    per_tmp = os.path.join(EMAIL_UPLOAD_DIR, f'per_{int(time.time()*1000)}_{f.filename}')
                    f.save(per_tmp)
                    try:
                        file_key = _email_upload_attachment(session, csrftoken, per_tmp, f.filename)
                        if file_key:
                            per_keys.append(file_key)
                            try:
                                f.filename.encode('ascii')
                                safe_name = f.filename
                            except UnicodeEncodeError:
                                b64 = base64.b64encode(f.filename.encode('utf-8')).decode('ascii')
                                safe_name = f'=?UTF-8?B?{b64}?='
                            per_names.append(safe_name)
                    except Exception:
                        pass
                    finally:
                        try: os.remove(per_tmp)
                        except OSError: pass
            all_att_keys = common_keys + per_keys
            all_att_names = common_names + per_names
            att_list_str = ','.join(all_att_keys) if all_att_keys else ''
            att_name_str = ','.join(all_att_names) if all_att_names else ''
            send_data = {
                'from': MAIL_CONFIG['username'], 'to': to_email, 'cc': ','.join(cc_emails),
                'bcc': '', 'fast': '0', 'content': content_b64, 'contentType': '1',
                'subject': subject, 'attachmentList': att_list_str,
                'attachmentNameList': att_name_str, 'dnt': '0',
                'action': 'send', 'sendMode': '0', 'saveSended': '1',
                'securityDestroy': '0', 'acceptSmsphones': '', 'acceptSmsKey': '',
                'securityCode': security_code,
            }
            encoded_body = urllib.parse.urlencode(send_data, encoding='utf-8')
            send_resp = session.post('https://mail.chinatelecom.cn/w2/mail/sendMail',
                data=encoded_body.encode('utf-8'),
                headers={'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'}, timeout=30)
            if send_resp.status_code == 200:
                result = send_resp.json()
                if result.get('code') == 0:
                    success_count += 1
                else:
                    fail_list.append(f"{to_email}: {result.get('desc', '发送失败')}")
            else:
                fail_list.append(f"{to_email}: HTTP {send_resp.status_code}")
        msg = f'成功发送 {success_count}/{len(items)} 封邮件'
        if fail_list:
            msg += f'，失败详情: {"; ".join(fail_list[:5])}'
        for tmp_path, _ in common_tmp_files:
            try: os.remove(tmp_path)
            except OSError: pass
        return jsonify({'success': success_count > 0, 'message': msg,
                        'success_count': success_count, 'fail_count': len(fail_list)})
    except req.exceptions.ConnectionError:
        return jsonify({'success': False, 'message': '网络连接失败'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': f'批量发送失败: {str(e)}'}), 500


# ---- 黑箱登录 ----
@app.route('/api/email/login/status', methods=['GET'])
def email_get_login_status():
    return jsonify(login_state)


@app.route('/api/email/login/start', methods=['POST'])
def email_start_login():
    global login_state, MAIL_CONFIG
    if login_state.get('status') in ('logging_in', 'waiting_code', 'verifying'):
        return jsonify({'success': False, 'message': '登录流程进行中，请稍候'})
    data = request.json or {}
    account = data.get('account', '').strip()
    password = data.get('password', '').strip()
    phone = data.get('phone', '').strip()
    if account: MAIL_CONFIG['account'] = account
    if password: MAIL_CONFIG['password'] = password
    if phone: MAIL_CONFIG['phone'] = phone
    _save_login_creds(MAIL_CONFIG['account'], MAIL_CONFIG['password'], MAIL_CONFIG['phone'])
    login_state = {'status': 'logging_in', 'message': '正在启动浏览器自动登录...', 'code': None, 'browser_open': False}
    thread = threading.Thread(target=_blackbox_login_worker, daemon=True)
    thread.start()
    return jsonify({'success': True, 'message': '登录流程已启动'})


@app.route('/api/email/login/creds', methods=['GET'])
def email_get_login_creds():
    creds = _load_login_creds()
    phone = creds.get('phone', '')
    phone_display = phone[:3] + '****' + phone[-4:] if len(phone) >= 7 else phone
    return jsonify({'success': True, 'data': {
        'account': creds.get('account', ''), 'password': creds.get('password', ''),
        'phone': phone, 'phone_display': phone_display}})


@app.route('/api/email/logout', methods=['POST'])
def email_logout():
    global login_state
    login_state = {'status': 'idle', 'message': '已退出登录', 'code': None, 'browser_open': False}
    if os.path.exists(EMAIL_COOKIES_FILE): os.remove(EMAIL_COOKIES_FILE)
    csrftoken_file = os.path.join(EMAIL_DATA_DIR, 'csrftoken.txt')
    if os.path.exists(csrftoken_file): os.remove(csrftoken_file)
    return jsonify({'success': True, 'message': '已退出登录'})


@app.route('/api/email/login/verify', methods=['POST'])
def email_submit_verify_code():
    global login_state
    if login_state.get('status') != 'waiting_code':
        return jsonify({'success': False, 'message': '当前不需要验证码'})
    code = request.json.get('code', '')
    if not code or len(code) < 4:
        return jsonify({'success': False, 'message': '验证码格式不正确'})
    login_state['code'] = code
    login_state['status'] = 'verifying'
    login_state['message'] = '正在提交验证码...'
    return jsonify({'success': True, 'message': '验证码已提交'})


@app.route('/api/email/login/cancel', methods=['POST'])
def email_cancel_login():
    global login_state
    login_state = {'status': 'idle', 'message': '登录已取消', 'code': None, 'browser_open': False}
    return jsonify({'success': True})


def _blackbox_login_worker():
    global login_state
    browser = None
    try:
        from playwright.sync_api import sync_playwright
        login_state['message'] = '正在启动浏览器...'
        chrome_path = r'C:\Program Files\Google\Chrome\Application\chrome.exe'
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, executable_path=chrome_path,
                args=['--no-sandbox', '--disable-blink-features=AutomationControlled'])
            context = browser.new_context(viewport={'width': 1280, 'height': 800},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
            page = context.new_page()
            login_state['message'] = '正在打开邮箱登录页面...'
            page.goto('https://mail.chinatelecom.cn/mail/index.html#/user/login', wait_until='networkidle', timeout=30000)
            page.wait_for_timeout(3000)
            login_state['message'] = '正在填入账号...'
            account_input = page.get_by_placeholder('邮箱账号/管理员账号')
            account_input.wait_for(state='visible', timeout=10000)
            account_input.click(); page.wait_for_timeout(300)
            account_input.fill(MAIL_CONFIG['account']); page.wait_for_timeout(500)
            login_state['message'] = '正在填入密码...'
            pwd_input = page.get_by_placeholder('输入邮箱密码')
            pwd_input.wait_for(state='visible', timeout=5000)
            pwd_input.click(); page.wait_for_timeout(200)
            pwd_input.fill(MAIL_CONFIG['password']); page.wait_for_timeout(500)
            login_state['message'] = '正在点击登录...'
            login_btn = page.get_by_role('button', name='登 录')
            login_btn.wait_for(state='visible', timeout=5000)
            login_btn.click()
            login_state['message'] = '等待二次验证页面...'
            page.wait_for_url('**/user/auth**', timeout=15000)
            page.wait_for_timeout(2000)
            login_state['message'] = '正在切换到手机验证码验证...'
            sms_radio = page.get_by_role('radio', name='手机验证码验证')
            sms_radio.wait_for(state='visible', timeout=10000)
            sms_radio.click(); page.wait_for_timeout(1500)
            login_state['message'] = '正在获取验证码...'
            get_code_btn = page.get_by_role('button', name='获取验证码')
            get_code_btn.wait_for(state='visible', timeout=10000)
            get_code_btn.click(); page.wait_for_timeout(2000)
            login_state['status'] = 'waiting_code'
            login_state['message'] = '验证码已发送到手机，请在下方输入'
            wait_start = time.time()
            while login_state.get('status') == 'waiting_code':
                page.wait_for_timeout(500)
                elapsed = time.time() - wait_start
                if int(elapsed) % 20 == 0 and int(elapsed) > 0:
                    try: page.mouse.move(100, 100); page.mouse.move(200, 200)
                    except Exception: pass
                if elapsed > 300:
                    login_state['status'] = 'failed'
                    login_state['message'] = '等待验证码超时（5分钟）'
                    return
            if login_state.get('status') != 'verifying': return
            code = login_state.get('code', '')
            login_state['message'] = '正在填入验证码...'
            try:
                code_input = page.locator('input[placeholder="请输入验证码"]')
                code_input.wait_for(state='visible', timeout=15000)
                code_input.click(); page.wait_for_timeout(200)
                code_input.fill(''); code_input.fill(code)
                page.wait_for_timeout(500)
            except Exception as e:
                login_state['status'] = 'failed'
                login_state['message'] = f'未找到验证码输入框: {str(e)}'
                return
            login_state['message'] = '正在确认登录...'
            clicked = False
            for sel in ['button.confirm-btn', 'button.ant-btn-primary.confirm-btn']:
                try:
                    btn = page.locator(sel)
                    if btn.count() > 0 and btn.first.is_visible(timeout=3000):
                        btn.first.click(); clicked = True; break
                except Exception: continue
            if not clicked:
                page.wait_for_timeout(3000)
                current_url = page.url
                if 'login' not in current_url.lower() and 'auth' not in current_url.lower():
                    clicked = True
            if not clicked:
                login_state['status'] = 'failed'
                login_state['message'] = '登录确认失败，请重试'
                return
            page.wait_for_timeout(5000)
            current_url = page.url
            login_success = False
            if 'login' not in current_url.lower() and 'auth' not in current_url.lower():
                login_success = True
            else:
                try:
                    if page.locator('text=退出').count() > 0 or page.locator('text=资源管理').count() > 0:
                        login_success = True
                except Exception: pass
            if login_success:
                cookies = context.cookies()
                _email_save_json(EMAIL_COOKIES_FILE, cookies)
                try:
                    csrftoken = page.evaluate(
                        '() => { try { const d = JSON.parse(localStorage.getItem("N_W_C_T") || "{}"); '
                        'const keys = Object.keys(d).filter(k => d[k] === true); '
                        'return keys.length > 0 ? keys[keys.length - 1] : ""; } catch(e) { return ""; } }')
                    if csrftoken:
                        csrftoken = re.sub(r'[\u200b\u200c\u200d\ufeff\u00a0\s]', '', csrftoken)
                    if csrftoken:
                        with open(os.path.join(EMAIL_DATA_DIR, 'csrftoken.txt'), 'w') as f:
                            f.write(csrftoken)
                except Exception: pass
                login_state['status'] = 'success'
                login_state['message'] = '登录成功！可以正常发送邮件了。'
            else:
                login_state['status'] = 'failed'
                login_state['message'] = '登录失败，请重试。'
            login_state['browser_open'] = False
            try: browser.close()
            except Exception: pass
    except ImportError:
        login_state['status'] = 'failed'
        login_state['message'] = 'Playwright未正确安装'
    except Exception as e:
        login_state['status'] = 'failed'
        login_state['message'] = f'登录过程出错: {str(e)}'
        login_state['browser_open'] = False
        try:
            if browser: browser.close()
        except Exception: pass


# ---- 邮箱配置 ----
@app.route('/api/email/mail-config', methods=['GET'])
def email_get_mail_config():
    return jsonify({'success': True, 'data': {
        'smtp_server': 'smtp.chinatelecom.cn', 'smtp_port': 587,
        'pop3_server': 'pop.chinatelecom.cn', 'pop3_port': 995,
        'imap_server': 'imap.chinatelecom.cn', 'imap_port': 993,
        'username': MAIL_CONFIG['username'],
        'send_method': 'webapi'}})


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5557, debug=True)
