import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import shutil

def clean_name(name):
    """清洗姓名，去除括号内的内容（如地区），只保留核心姓名"""
    if not isinstance(name, str):
        return ""
    # 去除括号及括号内的内容，例如 "肖智宇(高新区)" -> "肖智宇"
    name = re.sub(r'\([^)]*\)', '', name)
    name = re.sub(r'\（[^)]*\）', '', name) # 全角括号
    return name.strip()

def copy_row_with_format(source_sheet, target_sheet, source_row, target_row, max_col):
    """复制行并保持格式"""
    for col in range(1, max_col + 1):
        source_cell = source_sheet.cell(row=source_row, column=col)
        target_cell = target_sheet.cell(row=target_row, column=col)

        # 复制值
        target_cell.value = source_cell.value

        # 复制格式
        if source_cell.has_style:
            target_cell.font = source_cell.font.copy()
            target_cell.border = source_cell.border.copy()
            target_cell.fill = source_cell.fill.copy()
            target_cell.number_format = source_cell.number_format
            target_cell.protection = source_cell.protection.copy()
            target_cell.alignment = source_cell.alignment.copy()

        # 复制列宽
        col_letter = get_column_letter(col)
        target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width

def copy_sheet_with_format(source_sheet, target_sheet, row_indices):
    """复制工作表并保持格式，只包含指定的行"""
    max_col = source_sheet.max_column
    max_row = source_sheet.max_row

    # 复制列宽
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width

    # 复制行高和行格式
    target_row = 1
    for source_row in row_indices:
        copy_row_with_format(source_sheet, target_sheet, source_row, target_row, max_col)

        # 复制行高
        if source_sheet.row_dimensions[source_row].height:
            target_sheet.row_dimensions[target_row].height = source_sheet.row_dimensions[source_row].height

        target_row += 1

    # 复制工作表级别的设置
    target_sheet.sheet_format = source_sheet.sheet_format
    target_sheet.page_setup = source_sheet.page_setup
    target_sheet.page_margins = source_sheet.page_margins

def main():
    # 1. 定义分局与客户经理的映射关系 (根据您提供的文本整理)
    # 格式: {分局名称: [姓名1, 姓名2...]}
    bureau_mapping = {
        "工业能源政企分局": ["朱晨静", "肖智宇", "杜云帆", "屈容", "陈谦", "唐璐", "杜秋宇"],
        "国有平台政企分局": ["付登会", "蓝旭辉", "田小兰", "颜小琳", "易琴", "杨文", "王越"],
        "健康医疗政企分局": ["潘昱忻", "黎明", "先有为", "王虹丹"],
        "金融证券政企分局": ["贺贤珍", "曹岚", "梁曦宇", "杨彭萱", "张静", "郑文", "刘浩林", "杨真", "杨玮萍"],
        "软件科研政企分局": ["雷世豪", "陈燕", "杨璨宇", "李春江", "潘邓浩", "蒋尧莉", "李永婷", "蒲竑佚", "周建龙", "杨楚琪"],
        "新经济政企分局": ["曾理", "韩思萌", "罗维", "罗霞", "杨佩东", "刘文博", "戚新鹏"],
        "政法应急政企分局": ["王万辞", "徐杨", "袁进", "朱林", "袁新博", "张皓秋", "林佳俊", "黄天玉"],
        "政务政企分局": ["薛笑枫", "廖晓东", "钱宇煊", "吴倩", "杜成思", "黄振宇"],
        "高新孵化园智改数转服务局": ["董小凤", "邱国锋", "龙俊儒"],
        "高新天府生命科技园智改数转服务局": ["杨佳", "蒋建国", "王斯祺"],
        "高新天府软件园智改数转服务局": ["谢思宇", "黄微", "李选玉", "朱勇", "谢勇", "何琼", "高欢", "周莉", "李艺"],
        # "金融城商客分局": ["姚尧", "王淑惠", "陈伟智", "曾宇嘉", "罗中伟", "张成铭", "刘荣"],
        # "新川商客分局": ["靳扬", "姜春阳", "郑黎霞", "杨晋"],
        # "天府新谷商客分局": ["李思锐", "彭倩", "黄燕", "刘鹏洋", "刘星月"],
        # "新会展商客分局": ["何艳", "樊志林", "周滨", "蒋稚薇"],
        # "天府国际商客分局": ["叶江", "蒋天佑", "杨茜", "李巧巧", "王辰雨", "廖华", "曾明全", "巫婷婷"],
        # "环球商客分局": ["曾明", "许可", "钟小燕", "肖福洋", "冯麟霞", "王琴丽"],
        # "大源商客分局": ["邱浩锋", "冯兰越", "冯特峰", "裴嘉轩", "何亚琪"],
        # "肖芳商客分局": ["任登科", "贾小东", "梁润", "陈雪"],
        # "府城商客分局": ["陈磊", "李若玉", "张小龙", "孙雯"],
        # "连锁商客分局": ["杨凤翥", "肖帆", "赵娇", "高毛茅", "温有军"],
        # "西信商客分局": ["杨力", "王宇", "任少杰", "周雨晴", "刘祖源", "胡文瀚", "赵川川"],
        # "东苑商客分局": ["雷蕾", "吴文宪", "孙艺丹", "张宇魁", "聂海林", "陈健明"]
    }
    # bureau_mapping = {
    #     "金融城商客分局": ["姚尧", "王淑惠", "陈伟智", "曾宇嘉", "罗中伟", "张成铭", "刘荣"],
    #     "新川商客分局": ["靳扬", "姜春阳", "郑黎霞", "杨晋"],
    #     "天府新谷商客分局": ["李思锐", "彭倩", "黄燕", "刘鹏洋", "刘星月"],
    #     "新会展商客分局": ["何艳", "樊志林", "周滨", "蒋稚薇"],
    #     "天府国际商客分局": ["叶江", "蒋天佑", "杨茜", "李巧巧", "王辰雨", "廖华", "曾明全", "巫婷婷"],
    #     "环球商客分局": ["曾明", "许可", "钟小燕", "肖福洋", "冯麟霞", "王琴丽"],
    #     "大源商客分局": ["邱浩锋", "冯兰越", "冯特峰", "裴嘉轩", "何亚琪"],
    #     "肖芳商客分局": ["任登科", "贾小东", "梁润", "陈雪"],
    #     "府城商客分局": ["陈磊", "李若玉", "张小龙", "孙雯"],
    #     "连锁商客分局": ["杨凤翥", "肖帆", "赵娇", "高毛茅", "温有军"],
    #     "西信商客分局": ["杨力", "王宇", "任少杰", "周雨晴", "刘祖源", "胡文瀚", "赵川川"],
    #     "东苑商客分局": ["雷蕾", "吴文宪", "孙艺丹", "张宇魁", "聂海林", "陈健明"]
    # }    

    # 2. 读取Excel文件路径
    file_path = "H:\电信相关的\商机通报\商机宽表26416-517\商机宽表 (1).xlsx"  # 请确保文件路径正确

    if not os.path.exists(file_path):
        print(f"错误：找不到文件 {file_path}，请检查文件名和路径。")
        return

    # 获取源文件所在目录和当前日期
    source_dir = os.path.dirname(file_path)
    current_date = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 创建输出文件夹
    output_folder = os.path.join(source_dir, f"分局拆分结果_{current_date}")
    os.makedirs(output_folder, exist_ok=True)

    # 3. 使用openpyxl读取原始Excel文件以保持格式
    try:
        source_wb = load_workbook(file_path)
        source_sheet = source_wb.active
    except Exception as e:
        print(f"读取文件时出错：{e}")
        return

    # 4. 使用pandas读取数据用于匹配
    try:
        df = pd.read_excel(file_path, sheet_name=source_sheet.title, dtype=str)
    except Exception as e:
        print(f"读取数据时出错：{e}")
        return

    # 5. 数据清洗与匹配
    # 创建一个字典来存储每个分局的行索引
    bureau_rows = {bureau: [] for bureau in bureau_mapping.keys()}
    # 创建列表存储未匹配的行索引
    unmatched_rows = []

    # 首先添加表头（第1行）
    header_row = 1

    matched_count = 0
    unmatched_count = 0

    for index in range(len(df)):
        excel_row = index + 2  # Excel行号（第1行是表头，数据从第2行开始）

        # 获取客户经理列的值
        if '客户经理' in df.columns:
            manager_name_raw = df.iloc[index]['客户经理']
        elif 'AB1_客户经理' in df.columns:
            manager_name_raw = df.iloc[index]['AB1_客户经理']
        else:
            print(f"警告：找不到客户经理列，跳过第 {excel_row} 行")
            unmatched_rows.append(excel_row)
            unmatched_count += 1
            continue

        manager_name_clean = clean_name(manager_name_raw)

        matched = False
        for bureau, managers in bureau_mapping.items():
            if manager_name_clean in managers:
                bureau_rows[bureau].append(excel_row)
                matched_count += 1
                matched = True
                break

        if not matched:
            # 将未匹配的行存入列表
            unmatched_rows.append(excel_row)
            unmatched_count += 1
            print(f"警告：第 {excel_row} 行，客户经理 '{manager_name_raw}' 未在分局对照表中找到。")

    print(f"\n数据匹配完成！")
    print(f"成功匹配行数：{matched_count}")
    print(f"未匹配行数：{unmatched_count}")

    # 6. 为每个分局创建单独的Excel文件（命名格式：分局名称_年月日.xlsx）
    for bureau_name, row_indices in bureau_rows.items():
        if row_indices:  # 如果该分局有数据
            # 创建新的工作簿
            from openpyxl import Workbook
            target_wb = Workbook()
            target_sheet = target_wb.active

            # 复制表头和数据行
            all_rows = [header_row] + row_indices
            copy_sheet_with_format(source_sheet, target_sheet, all_rows)

            # 创建安全的文件名
            safe_name = bureau_name.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')
            output_file = os.path.join(output_folder, f"{safe_name}_{current_date}.xlsx")
            target_wb.save(output_file)
            print(f"已生成分局文件：{output_file} ({len(row_indices)} 行)")
        else:
            print(f"分局 '{bureau_name}' 没有匹配的数据，跳过生成独立文件。")

    # 7. 创建汇总Excel文件（包含所有分局的工作表）
    summary_file = os.path.join(output_folder, f"行业商机数据汇总_{current_date}.xlsx")
    from openpyxl import Workbook
    summary_wb = Workbook()

    # 删除默认的Sheet
    if 'Sheet' in summary_wb.sheetnames:
        del summary_wb['Sheet']

    for bureau_name, row_indices in bureau_rows.items():
        if row_indices:  # 如果该分局有数据
            # 创建新的工作表
            sheet_name = bureau_name[:31]  # Excel限制31字符
            target_sheet = summary_wb.create_sheet(title=sheet_name)

            # 复制表头和数据行
            all_rows = [header_row] + row_indices
            copy_sheet_with_format(source_sheet, target_sheet, all_rows)
        else:
            # 创建空工作表
            sheet_name = bureau_name[:31]
            summary_wb.create_sheet(title=sheet_name)

    summary_wb.save(summary_file)
    print(f"\n已生成汇总文件：{summary_file}")

    # 8. 为未匹配的数据创建单独的Excel文件
    if unmatched_rows:
        from openpyxl import Workbook
        unmatched_wb = Workbook()
        unmatched_sheet = unmatched_wb.active

        # 复制表头和未匹配的行
        all_rows = [header_row] + unmatched_rows
        copy_sheet_with_format(source_sheet, unmatched_sheet, all_rows)

        unmatched_file = os.path.join(output_folder, f"未匹配名单_{current_date}.xlsx")
        unmatched_wb.save(unmatched_file)
        print(f"已生成未匹配名单文件：{unmatched_file} ({len(unmatched_rows)} 行)")
    else:
        print(f"所有数据都已匹配，无需生成未匹配名单文件。")

    print(f"\n所有文件已保存到文件夹：{output_folder}")

if __name__ == "__main__":
    main()
